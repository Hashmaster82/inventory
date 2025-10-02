import json
import os
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
from datetime import datetime
from fpdf import FPDF
import webbrowser
import tkinter.font as tkFont
import sys
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional

# ----------------- Настройка логирования -----------------
logger = logging.getLogger("InventoryApp")
logger.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler = logging.FileHandler("inventory_app.log", encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(formatter)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(formatter)
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# ----------------- Вспомогательные функции -----------------
def _get_asset_path(filename: str) -> str:
    """Универсальный метод поиска asset-файлов (для шрифтов и т.п.)."""
    if getattr(sys, 'frozen', False):
        base_path = Path(sys._MEIPASS)
    else:
        base_path = Path(__file__).parent
    candidates = [
        base_path / 'assets' / 'fonts' / filename,
        base_path / filename
    ]
    for path in candidates:
        if path.exists():
            return str(path)
    raise FileNotFoundError(f"Файл не найден: {filename} в путях: {[str(p) for p in candidates]}")

def _safe_save_json(data: Any, filepath: Path) -> bool:
    """Сохраняет данные в JSON с использованием временного файла."""
    try:
        temp_file = filepath.with_suffix(filepath.suffix + ".tmp")
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        temp_file.replace(filepath)
        logger.info(f"Файл сохранён: {filepath}")
        return True
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить файл {filepath.name}: {e}")
        logger.error(f"Save {filepath}: {e}")
        return False

# ----------------- Класс PDF с поддержкой кириллицы и нумерацией страниц -----------------
class PDFWithCyrillic(FPDF):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        font_path = _get_asset_path('ChakraPetch-Regular.ttf')
        self.add_font('ChakraPetch', '', font_path, uni=True)
        self.alias_nb_pages()

    def footer(self):
        self.set_y(-15)
        self.set_font('ChakraPetch', '', 10)
        self.cell(0, 10, f'Страница {self.page_no()} из {{nb}}', 0, 0, 'C')

# ----------------- Основной класс приложения -----------------
class InventoryApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Система инвентаризации оборудования")
        self.root.state('zoomed')
        # === Установка иконки приложения ===
        try:
            icon_path = Path(__file__).parent / "app.ico"
            if icon_path.exists():
                self.root.iconbitmap(str(icon_path))
            else:
                logger.warning("Файл иконки app.ico не найден")
        except Exception as e:
            logger.error(f"Не удалось установить иконку: {e}")

        self.default_font = tkFont.Font(family='Arial', size=14)
        self.root.option_add("*Font", self.default_font)

        # === Настройки: загрузка или выбор каталога данных ===
        self.settings_file = Path(__file__).parent / "settings.json"
        self.data_dir = self.load_settings()
        if not self.data_dir:
            self.data_dir = self.choose_data_directory_on_start()
            if not self.data_dir:
                messagebox.showerror("Ошибка", "Не выбран каталог данных. Программа будет закрыта.")
                root.quit()
                return
            self.save_settings(self.data_dir)

        # === Пути к файлам (все в self.data_dir) ===
        self.inventory_file = self.data_dir / "inventory.json"
        self.equipment_types_file = self.data_dir / "equipment_types.json"
        self.history_file = self.data_dir / "history.json"
        self.employees_file = self.data_dir / "sotrudniki.json"
        self.data_dir.mkdir(exist_ok=True)

        self.inventory_data = self.load_data()
        self.equipment_types = self.load_equipment_types()
        self.history_data = self.load_history()
        self.employees_list = self.load_employees()

        self.unsaved_changes = False
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # === Для редактирования записи ===
        self.current_edit_index = None
        self.edit_entries = {}
        self.create_widgets()
        self.update_window_title()
        self.auto_save_interval = 300000  # 5 минут
        self.schedule_auto_save()

        # Клавиатурные сокращения
        self.root.bind('<Control-s>', lambda e: self.save_data() and self.mark_saved())
        self.root.bind('<Control-f>', lambda e: self.search_entry.focus_set())
        self.root.bind('<Delete>', self.delete_selected_item)

        # Статусная строка
        self.status_var = tk.StringVar(value="Готово")
        self.status_bar = ttk.Label(root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def update_window_title(self):
        base_title = "Система инвентаризации оборудования"
        suffix = "*" if self.unsaved_changes else ""
        self.root.title(base_title + suffix)

    def mark_saved(self):
        self.unsaved_changes = False
        self.update_window_title()

    def on_closing(self):
        if self.unsaved_changes:
            answer = messagebox.askyesnocancel(
                "Несохранённые изменения",
                "Есть несохранённые изменения. Сохранить перед выходом?"
            )
            if answer is True:
                if self.save_data():
                    self.root.destroy()
            elif answer is False:
                self.root.destroy()
        else:
            self.root.destroy()

    def is_serial_number_unique(self, serial_number: str, exclude_index: Optional[int] = None) -> bool:
        for i, item in enumerate(self.inventory_data):
            if exclude_index is not None and i == exclude_index:
                continue
            if item.get('serial_number') == serial_number:
                return False
        return True

    # =============== РАБОТА С НАСТРОЙКАМИ ===============
    def load_settings(self) -> Optional[Path]:
        try:
            if self.settings_file.exists():
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    data_dir = settings.get("data_directory")
                    if data_dir and Path(data_dir).is_dir():
                        return Path(data_dir)
        except Exception as e:
            logger.error(f"Load settings: {e}")
        return None

    def save_settings(self, data_dir: Path):
        try:
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump({"data_directory": str(data_dir)}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить настройки: {e}")
            logger.error(f"Save settings: {e}")

    def choose_data_directory_on_start(self) -> Optional[Path]:
        messagebox.showinfo("Первый запуск", "Пожалуйста, выберите каталог для хранения данных инвентаризации.")
        directory = filedialog.askdirectory(title="Выберите каталог для хранения данных")
        return Path(directory) if directory else None

    # =============== РАБОТА С ТИПАМИ ОБОРУДОВАНИЯ ===============
    def load_equipment_types(self) -> List[str]:
        try:
            if self.equipment_types_file.exists():
                with open(self.equipment_types_file, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    return data if isinstance(data, list) else []
            else:
                default_types = ["Монитор", "Сисблок", "МФУ", "Клавиатура", "Мышь", "Наушники"]
                self.save_equipment_types(default_types)
                return default_types
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить типы оборудования: {e}")
            logger.error(f"Load equipment types: {e}")
            return []

    def save_equipment_types(self, types_list: List[str]) -> bool:
        return _safe_save_json(types_list, self.equipment_types_file)

    # =============== РАБОТА СО СПИСКОМ СОТРУДНИКОВ ===============
    def load_employees(self) -> List[str]:
        try:
            if self.employees_file.exists():
                with open(self.employees_file, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    return data if isinstance(data, list) else []
            else:
                self.save_employees([])
                return []
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить список сотрудников: {e}")
            logger.error(f"Load employees: {e}")
            return []

    def save_employees(self, employees_list: List[str]) -> bool:
        return _safe_save_json(employees_list, self.employees_file)

    def add_employee(self, employee_name: str) -> bool:
        if not employee_name.strip():
            return False
        employee_name = employee_name.strip()
        if employee_name in self.employees_list:
            messagebox.showwarning("Предупреждение", "Этот сотрудник уже существует.")
            return False
        self.employees_list.append(employee_name)
        if self.save_employees(self.employees_list):
            self.unsaved_changes = True
            self.update_window_title()
            return True
        return False

    def delete_employee(self, employee_name: str) -> bool:
        if not employee_name:
            return False
        in_use = any(item.get('assignment', '') == employee_name for item in self.inventory_data)
        if in_use:
            messagebox.showerror("Ошибка", f"Сотрудник '{employee_name}' используется в записях. Удаление запрещено.")
            return False
        if messagebox.askyesno("Подтверждение", f"Удалить сотрудника '{employee_name}'?"):
            self.employees_list.remove(employee_name)
            if self.save_employees(self.employees_list):
                self.unsaved_changes = True
                self.update_window_title()
                return True
        return False

    def update_employee_comboboxes(self):
        if hasattr(self, 'assignment_combo'):
            self.assignment_combo['values'] = [""] + sorted(self.employees_list)
            if self.employees_list:
                self.assignment_var.set(self.employees_list[0])
            else:
                self.assignment_var.set('')
        if hasattr(self, 'employee_combo'):
            self.employee_combo['values'] = [""] + sorted(self.employees_list)
            if self.employees_list:
                self.employee_var.set(self.employees_list[0])
            else:
                self.employee_var.set('')
        if hasattr(self, 'history_employee_combo'):
            self.history_employee_combo['values'] = [""] + sorted(set(
                item.get('assignment', '') for item in self.inventory_data if item.get('assignment')
            ))

    # =============== РАБОТА С ИСТОРИЕЙ ===============
    def load_history(self) -> Dict[str, List[Dict[str, str]]]:
        try:
            if self.history_file.exists():
                with open(self.history_file, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    return data if isinstance(data, dict) else {}
            else:
                return {}
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить историю: {e}")
            logger.error(f"Load history: {e}")
            return {}

    def save_history(self) -> bool:
        return _safe_save_json(self.history_data, self.history_file)

    def add_to_history(self, serial_number: str, assignment: str, date: str):
        if not serial_number or not assignment:
            return
        if serial_number not in self.history_data:
            self.history_data[serial_number] = []
        entry = {"assignment": assignment, "date": date}
        if entry not in self.history_data[serial_number]:
            self.history_data[serial_number].append(entry)
            self.save_history()

    def get_history_for_equipment(self, serial_number: str) -> List[Dict[str, str]]:
        return self.history_data.get(serial_number, [])

    def initialize_history_from_inventory(self):
        """Заполняет history.json начальными записями из inventory.json."""
        if not self.inventory_data:
            messagebox.showinfo("Информация", "Нет данных в инвентаризации для инициализации истории.")
            return
        updated = False
        for item in self.inventory_data:
            serial = item.get('serial_number')
            assignment = item.get('assignment')
            date = item.get('date')
            if not serial or not assignment or not date:
                continue
            if serial not in self.history_data:
                self.history_data[serial] = []
            existing = any(rec.get('assignment') == assignment and rec.get('date') == date for rec in self.history_data[serial])
            if not existing:
                self.history_data[serial].append({"assignment": assignment, "date": date})
                updated = True
        if updated:
            if self.save_history():
                messagebox.showinfo("Успех", "История успешно инициализирована из текущих данных инвентаризации.")
                self.show_full_history()
            else:
                messagebox.showerror("Ошибка", "Не удалось сохранить историю.")
        else:
            messagebox.showinfo("Информация", "История уже содержит все начальные записи.")

    # =============== ВСПОМОГАТЕЛЬНЫЙ МЕТОД ДЛЯ ПОЛУЧЕНИЯ МОДЕЛИ ===============
    def _get_model_by_serial(self, serial: str) -> str:
        """Вспомогательная функция: получает модель по серийному номеру."""
        for item in self.inventory_data:
            if item.get('serial_number') == serial:
                return item.get('model', '-')
        return "-"

    # =============== ОСНОВНАЯ ЛОГИКА ===============
    def load_data(self) -> List[Dict[str, Any]]:
        try:
            if self.inventory_file.exists():
                with open(self.inventory_file, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    if not isinstance(data, list):
                        raise ValueError("Файл должен содержать массив объектов")
                    return data
            else:
                with open(self.inventory_file, 'w', encoding='utf-8') as file:
                    json.dump([], file)
                return []
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить данные: {e}")
            logger.error(f"Load inventory data: {e}")
            return []

    def save_data(self) -> bool:
        if _safe_save_json(self.inventory_data, self.inventory_file):
            self.mark_saved()
            return True
        return False

    def create_backup(self):
        try:
            backup_dir = self.data_dir / "backups"
            backup_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            files_to_backup = [
                self.inventory_file,
                self.history_file,
                self.employees_file,
                self.equipment_types_file
            ]
            backed_up = []
            for src_file in files_to_backup:
                if src_file.exists():
                    backup_name = src_file.name.replace(".json", f"_backup_{timestamp}.json")
                    dst_file = backup_dir / backup_name
                    import shutil
                    shutil.copy2(src_file, dst_file)
                    backed_up.append(dst_file)
                else:
                    logger.info(f"Файл для бэкапа не найден: {src_file}")
            all_backups = sorted(backup_dir.glob("*.json"), key=os.path.getmtime, reverse=True)
            for old_backup in all_backups[10:]:
                old_backup.unlink()
                logger.info(f"Удалён старый бэкап: {old_backup}")
            if backed_up:
                self.status_var.set("Резервные копии созданы")
                messagebox.showinfo("Успех", f"Резервные копии созданы:\n" + "\n".join(map(str, backed_up)))
            else:
                self.status_var.set("Нет файлов для резервного копирования")
                messagebox.showwarning("Предупреждение", "Нет файлов для резервного копирования")
        except Exception as e:
            self.status_var.set("Ошибка при создании резервной копии")
            messagebox.showerror("Ошибка", f"Не удалось создать резервную копию: {e}")
            logger.error(f"Create backup: {e}")

    def bind_clipboard_events(self, widget):
        def do_copy(event): widget.event_generate("<<Copy>>"); return "break"
        def do_cut(event): widget.event_generate("<<Cut>>"); return "break"
        def do_paste(event): widget.event_generate("<<Paste>>"); return "break"
        widget.bind("<Control-c>", do_copy)
        widget.bind("<Control-x>", do_cut)
        widget.bind("<Control-v>", do_paste)

    def create_widgets(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # ====== ВЕРХНЯЯ ПАНЕЛЬ КНОПОК ======
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10, fill='x')
        button_style = ttk.Style()
        button_style.configure('Small.TButton', font=('Arial', 12), padding=6)

        backup_button = ttk.Button(button_frame, text="📂 Создать резервную копию",
                                   command=self.create_backup, style='Small.TButton')
        backup_button.pack(side='left', padx=5, fill='x', expand=True)

        pdf_button = ttk.Button(button_frame, text="📄 Выгрузить полный отчет в PDF",
                                command=self.export_to_pdf, style='Small.TButton')
        pdf_button.pack(side='left', padx=5, fill='x', expand=True)

        excel_button = ttk.Button(button_frame, text="📊 Выгрузить полный отчёт в Excel",
                                  command=self.export_to_excel, style='Small.TButton')
        excel_button.pack(side='left', padx=5, fill='x', expand=True)

        graph_button = ttk.Button(button_frame, text="📈 График",
                                  command=self.show_equipment_graph, style='Small.TButton')
        graph_button.pack(side='left', padx=5, fill='x', expand=True)

        # ====== НАБОР ВКЛАДОК ======
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True)
        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 18, 'bold'))
        style.configure('TNotebook.Tab', font=('Arial', 16, 'bold'), padding=[20, 10])

        self.show_all_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.show_all_frame, text="Показать всё")

        self.add_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.add_frame, text="Добавить оборудование")

        self.search_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.search_frame, text="Поиск оборудования")

        self.employee_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.employee_frame, text="Сотрудники")

        self.equipment_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.equipment_frame, text="Оборудование")

        self.settings_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_frame, text="Настройки")

        self.history_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.history_frame, text="История")

        self.transfers_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.transfers_frame, text="🔄 Передачи")

        self.about_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.about_frame, text="Инфо")

        self.create_history_tab()
        self.create_employee_tab()
        self.create_show_all_tab()
        self.create_add_tab()
        self.create_search_tab()
        self.create_equipment_tab()
        self.create_settings_tab()
        self.create_about_tab()
        self.create_transfers_tab()

        self.notebook.select(self.show_all_frame)

    # =============== ВКЛАДКА: ПОКАЗАТЬ ВСЁ ===============
    def create_show_all_tab(self):
        refresh_button = ttk.Button(self.show_all_frame, text="Обновить данные",
                                    command=self.show_all_data, style='Small.TButton')
        refresh_button.pack(pady=5)

        table_frame = ttk.Frame(self.show_all_frame)
        table_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ("Тип", "Модель", "Серийный номер", "Закрепление", "Дата", "Комментарии")
        self.all_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=20)
        for col in columns:
            self.all_tree.heading(col, text=col,
                                  command=lambda _col=col: self.treeview_sort_column(self.all_tree, _col, False))
            self.all_tree.column(col, width=150, anchor='center')

        self.all_tree.bind('<Double-1>', self.on_tree_double_click)
        self.all_tree.bind('<Button-3>', self.show_context_menu)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.all_tree.yview)
        self.all_tree.configure(yscrollcommand=scrollbar.set)
        self.all_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # === КОНТЕКСТНОЕ МЕНЮ: "Редактировать" НАВЕРХ ===
        self.all_context_menu = tk.Menu(self.all_tree, tearoff=0)
        self.all_context_menu.add_command(label="✏️ Редактировать", command=self.edit_selected_item)
        self.all_context_menu.add_command(label="Передать", command=self.transfer_selected_item)
        self.all_context_menu.add_command(label="Удалить запись", command=self.delete_selected_item)

        self.show_all_data()

        # === БЛОК РЕДАКТИРОВАНИЯ ВНИЗУ ===
        edit_frame = ttk.LabelFrame(self.show_all_frame, text="Редактирование записи", padding=10)
        edit_frame.pack(fill='x', padx=10, pady=(10, 0))

        fields = [
            ("Тип оборудования", "equipment_type"),
            ("Модель", "model"),
            ("Серийный номер", "serial_number"),
            ("Закрепление", "assignment"),
            ("Дата", "date"),
            ("Комментарии", "comments")
        ]
        self.edit_entries = {}
        for i, (label_text, field_name) in enumerate(fields):
            label = ttk.Label(edit_frame, text=label_text + ":")
            label.grid(row=i, column=0, sticky='w', padx=5, pady=3)
            if field_name == "equipment_type":
                var = tk.StringVar()
                combo = ttk.Combobox(edit_frame, textvariable=var, values=sorted(self.equipment_types), width=30)
                combo.grid(row=i, column=1, padx=5, pady=3, sticky='we')
                self.edit_entries[field_name] = (var, combo)
            elif field_name == "assignment":
                var = tk.StringVar()
                combo = ttk.Combobox(edit_frame, textvariable=var, values=[""] + sorted(self.employees_list), width=30)
                combo.grid(row=i, column=1, padx=5, pady=3, sticky='we')
                self.edit_entries[field_name] = (var, combo)
            elif field_name == "comments":
                text = scrolledtext.ScrolledText(edit_frame, width=40, height=3, font=('Arial', 12))
                text.grid(row=i, column=1, padx=5, pady=3, sticky='we')
                self.edit_entries[field_name] = text
            else:
                entry = ttk.Entry(edit_frame, width=40)
                entry.grid(row=i, column=1, padx=5, pady=3, sticky='we')
                self.edit_entries[field_name] = entry

        # Кнопки
        btn_frame = ttk.Frame(edit_frame)
        btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=10)
        save_btn = ttk.Button(btn_frame, text="💾 Сохранить", command=self.save_edited_item, style='Small.TButton')
        save_btn.pack(side='left', padx=5)
        cancel_btn = ttk.Button(btn_frame, text="❌ Отмена", command=self.cancel_edit, style='Small.TButton')
        cancel_btn.pack(side='left', padx=5)

        edit_frame.columnconfigure(1, weight=1)

    # =============== МЕТОДЫ РЕДАКТИРОВАНИЯ ===============
    def edit_selected_item(self):
        selected = self.all_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите запись для редактирования")
            return
        values = self.all_tree.item(selected[0], 'values')
        serial = values[2]
        idx = next((i for i, item in enumerate(self.inventory_data) if item.get('serial_number') == serial), None)
        if idx is None:
            messagebox.showerror("Ошибка", "Запись не найдена в данных")
            return
        item = self.inventory_data[idx]
        self.current_edit_index = idx
        self.edit_entries['equipment_type'][0].set(item.get('equipment_type', ''))
        self.edit_entries['model'].delete(0, tk.END)
        self.edit_entries['model'].insert(0, item.get('model', ''))
        self.edit_entries['serial_number'].delete(0, tk.END)
        self.edit_entries['serial_number'].insert(0, item.get('serial_number', ''))
        self.edit_entries['assignment'][0].set(item.get('assignment', ''))
        self.edit_entries['date'].delete(0, tk.END)
        self.edit_entries['date'].insert(0, item.get('date', ''))
        if isinstance(self.edit_entries['comments'], scrolledtext.ScrolledText):
            self.edit_entries['comments'].delete('1.0', tk.END)
            self.edit_entries['comments'].insert('1.0', item.get('comments', ''))

    def save_edited_item(self):
        if self.current_edit_index is None:
            return

        item = self.inventory_data[self.current_edit_index]
        new_data = {}
        new_data['equipment_type'] = self.edit_entries['equipment_type'][0].get().strip()
        new_data['model'] = self.edit_entries['model'].get().strip()
        new_data['serial_number'] = self.edit_entries['serial_number'].get().strip()
        new_data['assignment'] = self.edit_entries['assignment'][0].get().strip()
        new_data['date'] = self.edit_entries['date'].get().strip()
        new_data['comments'] = self.edit_entries['comments'].get('1.0', tk.END).strip()

        if not new_data['equipment_type'] or not new_data['serial_number']:
            messagebox.showwarning("Ошибка", "Тип и серийный номер обязательны")
            return

        if new_data['equipment_type'] not in self.equipment_types:
            messagebox.showwarning("Ошибка", "Недопустимый тип оборудования")
            return

        if not self.is_serial_number_unique(new_data['serial_number'], exclude_index=self.current_edit_index):
            messagebox.showerror("Ошибка", f"Серийный номер '{new_data['serial_number']}' уже существует!")
            return

        try:
            datetime.strptime(new_data['date'], "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты. Используйте дд.мм.гггг")
            return

        old_assignment = item.get('assignment', '')
        item.update(new_data)

        if old_assignment != new_data['assignment'] and new_data['assignment']:
            self.add_to_history(new_data['serial_number'], new_data['assignment'], new_data['date'])

        # ИСПРАВЛЕНИЕ: НЕ ВЫЗЫВАЕМ show_all_data(), чтобы не перезагружать из файла!
        # Вместо этого обновим дерево вручную или просто сохраним
        if self.save_data():
            messagebox.showinfo("Успех", "Запись обновлена")
            self.cancel_edit()
            # Обновим интерфейс без перезагрузки из файла
            self.refresh_employee_list()
            self.update_history_combobox()
            self.update_serial_combobox()
            # Обновим строку в таблице вручную (опционально, но для точности)
            selected = self.all_tree.selection()
            if selected:
                self.all_tree.item(selected[0], values=(
                    new_data['equipment_type'],
                    new_data['model'],
                    new_data['serial_number'],
                    new_data['assignment'],
                    new_data['date'],
                    (new_data['comments'][:50] + '...') if len(new_data['comments']) > 50 else new_data['comments']
                ))

    def cancel_edit(self):
        self.current_edit_index = None
        for field, widget in self.edit_entries.items():
            if field in ('equipment_type', 'assignment'):
                widget[0].set('')
            elif field == 'comments':
                widget.delete('1.0', tk.END)
            else:
                widget.delete(0, tk.END)

    # =============== ВКЛАДКА: ПОИСК ОБОРУДОВАНИЯ ===============
    def create_search_tab(self):
        ttk.Label(self.search_frame, text="Поиск:", font=self.default_font).grid(row=0, column=0, sticky='w', padx=10,
                                                                                 pady=5)
        self.search_entry = ttk.Entry(self.search_frame, width=40, font=self.default_font)
        self.search_entry.grid(row=0, column=1, padx=10, pady=5, sticky='we')
        self.bind_clipboard_events(self.search_entry)
        self.search_entry.bind('<KeyRelease>', self.perform_search)

        ttk.Label(self.search_frame, text="Сотрудник:", font=self.default_font).grid(row=0, column=2, sticky='w',
                                                                                     padx=(20, 10), pady=5)
        self.search_employee_var = tk.StringVar()
        self.search_employee_combo = ttk.Combobox(self.search_frame, textvariable=self.search_employee_var,
                                                  values=[""] + sorted(self.employees_list), width=25,
                                                  font=self.default_font)
        self.search_employee_combo.grid(row=0, column=3, padx=10, pady=5, sticky='we')
        self.search_employee_combo.bind('<<ComboboxSelected>>', self.perform_search)

        clear_button = ttk.Button(self.search_frame, text="Очистить", command=self.clear_search, style='Small.TButton')
        clear_button.grid(row=0, column=4, padx=10, pady=5)

        columns = ("Тип", "Модель", "Серийный номер", "Закрепление", "Дата", "Комментарии")
        self.search_tree = ttk.Treeview(self.search_frame, columns=columns, show='headings', height=15)
        for col in columns:
            self.search_tree.heading(col, text=col,
                                     command=lambda _col=col: self.treeview_sort_column(self.search_tree, _col, False))
            self.search_tree.column(col, width=150, anchor='center')

        self.search_tree.bind('<Double-1>', self.on_tree_double_click)
        self.search_tree.bind('<Button-3>', self.show_context_menu)

        scrollbar = ttk.Scrollbar(self.search_frame, orient="vertical", command=self.search_tree.yview)
        self.search_tree.configure(yscrollcommand=scrollbar.set)
        self.search_tree.grid(row=1, column=0, columnspan=5, padx=10, pady=5, sticky='nsew')
        scrollbar.grid(row=1, column=5, sticky='ns', pady=5)

        export_pdf_btn = ttk.Button(self.search_frame, text="📄 Экспорт результатов поиска в PDF",
                                    command=self.export_search_results_to_pdf, style='Small.TButton')
        export_pdf_btn.grid(row=2, column=0, columnspan=5, pady=10, sticky='we')

        export_excel_btn = ttk.Button(self.search_frame, text="📊 Экспорт результатов поиска в Excel",
                                      command=self.export_search_results_to_excel, style='Small.TButton')
        export_excel_btn.grid(row=3, column=0, columnspan=5, pady=5, sticky='we')

        self.search_context_menu = tk.Menu(self.search_tree, tearoff=0)
        self.search_context_menu.add_command(label="Удалить запись", command=self.delete_selected_item)
        self.search_context_menu.add_command(label="Передать", command=self.transfer_selected_item)

        self.search_frame.columnconfigure(1, weight=1)
        self.search_frame.columnconfigure(3, weight=1)
        self.search_frame.rowconfigure(1, weight=1)

    # =============== ВКЛАДКА: СОТРУДНИКИ ===============
    def create_employee_tab(self):
        ttk.Label(self.employee_frame, text="Выберите сотрудника:", font=self.default_font).grid(row=0, column=0,
                                                                                                 sticky='w', padx=10,
                                                                                                 pady=5)
        self.employee_var = tk.StringVar()
        self.employee_combo = ttk.Combobox(self.employee_frame, textvariable=self.employee_var,
                                           values=[""] + sorted(self.employees_list), width=50, font=self.default_font)
        self.employee_combo.grid(row=0, column=1, columnspan=3, padx=10, pady=5, sticky='we')
        self.bind_clipboard_events(self.employee_combo)
        self.employee_combo.bind('<<ComboboxSelected>>', self.show_employee_equipment)

        ttk.Label(self.employee_frame, text="Поиск сотрудника:", font=self.default_font).grid(row=1, column=0,
                                                                                              sticky='w', padx=10,
                                                                                              pady=5)
        self.employee_search_var = tk.StringVar()
        self.employee_search_entry = ttk.Entry(self.employee_frame, textvariable=self.employee_search_var, width=50,
                                               font=self.default_font)
        self.employee_search_entry.grid(row=1, column=1, columnspan=3, padx=10, pady=5, sticky='we')
        self.employee_search_entry.bind('<KeyRelease>', self.filter_employees_by_search)

        btn_frame_employees = ttk.Frame(self.employee_frame)
        btn_frame_employees.grid(row=2, column=0, columnspan=4, padx=10, pady=10, sticky='we')

        add_employee_btn = ttk.Button(btn_frame_employees, text="➕ Добавить сотрудника",
                                      command=self.open_add_employee_dialog, style='Small.TButton')
        add_employee_btn.pack(side='left', padx=(0, 10))

        # === ИЗМЕНЕНИЕ: кнопка "Из базы" по правому краю + пояснение ===
        load_from_base_btn = ttk.Button(btn_frame_employees, text="📥 Из базы",
                                        command=self.load_employees_from_inventory, style='Small.TButton')
        load_from_base_btn.pack(side='right', padx=10)

        load_hint = ttk.Label(btn_frame_employees, text="Нажимаем только при первом запуске программы!", foreground="gray", font=('Arial', 10))
        load_hint.pack(side='right', padx=(0, 10))

        refresh_button = ttk.Button(btn_frame_employees, text="🔄 Обновить",
                                    command=self.refresh_employee_list, style='Small.TButton')
        refresh_button.pack(side='left', padx=10)

        # === ТАБЛИЦА ОБОРУДОВАНИЯ СОТРУДНИКА ===
        columns = ("Тип", "Модель", "Серийный номер", "Дата", "Комментарии")
        self.employee_tree = ttk.Treeview(self.employee_frame, columns=columns, show='headings', height=10)
        for col in columns:
            self.employee_tree.heading(col, text=col,
                                       command=lambda _col=col: self.treeview_sort_column(self.employee_tree, _col,
                                                                                          False))
            self.employee_tree.column(col, width=150, anchor='center')

        self.employee_tree.bind('<Double-1>', self.on_tree_double_click)
        self.employee_tree.bind('<Button-3>', self.show_context_menu)

        tree_scrollbar = ttk.Scrollbar(self.employee_frame, orient="vertical", command=self.employee_tree.yview)
        self.employee_tree.configure(yscrollcommand=tree_scrollbar.set)
        self.employee_tree.grid(row=3, column=0, columnspan=4, padx=10, pady=5, sticky='nsew')
        tree_scrollbar.grid(row=3, column=4, sticky='ns', pady=5)

        export_pdf_btn = ttk.Button(self.employee_frame, text="📄 Экспорт в PDF",
                                    command=self.export_employee_results_to_pdf, style='Small.TButton')
        export_pdf_btn.grid(row=4, column=0, columnspan=4, pady=10, sticky='we')

        export_excel_btn = ttk.Button(self.employee_frame, text="📊 Экспорт в Excel",
                                      command=self.export_employee_results_to_excel, style='Small.TButton')
        export_excel_btn.grid(row=5, column=0, columnspan=4, pady=5, sticky='we')

        # === СПИСОК ВСЕХ СОТРУДНИКОВ — ПЕРЕНЕСЁН ВНИЗ ===
        ttk.Label(self.employee_frame, text="Список всех сотрудников:", font=self.default_font).grid(row=6, column=0,
                                                                                                     sticky='w',
                                                                                                     padx=10,
                                                                                                     pady=(20, 5))
        self.all_employees_listbox = tk.Listbox(self.employee_frame, font=self.default_font, height=8,
                                                exportselection=False)
        self.all_employees_listbox.grid(row=7, column=0, columnspan=4, padx=10, pady=5, sticky='nsew')
        self.all_employees_listbox.bind('<Button-3>', self.show_employees_context_menu)

        emp_scrollbar = ttk.Scrollbar(self.employee_frame, orient="vertical", command=self.all_employees_listbox.yview)
        self.all_employees_listbox.configure(yscrollcommand=emp_scrollbar.set)
        emp_scrollbar.grid(row=7, column=4, sticky='ns', pady=5)

        export_employees_btn = ttk.Button(self.employee_frame, text="📤 Экспорт списка сотрудников в Excel",
                                          command=self.export_employees_to_excel, style='Small.TButton')
        export_employees_btn.grid(row=8, column=0, columnspan=4, pady=(10, 0), sticky='we')

        self.employees_context_menu = tk.Menu(self.all_employees_listbox, tearoff=0)
        self.employees_context_menu.add_command(label="✏️ Редактировать", command=self.edit_selected_employee)
        self.employees_context_menu.add_command(label="🗑️ Удалить", command=self.delete_selected_employee)

        self.employee_context_menu = tk.Menu(self.employee_tree, tearoff=0)
        self.employee_context_menu.add_command(label="Удалить запись", command=self.delete_selected_item)
        self.employee_context_menu.add_command(label="Передать", command=self.transfer_selected_item)

        self.employee_frame.columnconfigure(1, weight=1)
        self.employee_frame.rowconfigure(3, weight=1)
        self.employee_frame.rowconfigure(7, weight=0)

        self.refresh_employee_list()

    def show_employees_context_menu(self, event):
        idx = self.all_employees_listbox.nearest(event.y)
        if idx >= 0:
            self.all_employees_listbox.selection_clear(0, tk.END)
            self.all_employees_listbox.selection_set(idx)
            self.employees_context_menu.post(event.x_root, event.y_root)

    def edit_selected_employee(self):
        selection = self.all_employees_listbox.curselection()
        if not selection:
            return
        old_name = self.all_employees_listbox.get(selection[0])
        dialog = tk.Toplevel(self.root)
        dialog.title("Редактировать сотрудника")
        dialog.geometry("500x200")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Новое имя сотрудника:", font=self.default_font).pack(pady=(20, 10))
        entry = ttk.Entry(dialog, width=45, font=self.default_font)
        entry.insert(0, old_name)
        entry.pack(pady=10)
        entry.focus()

        def save_edit():
            new_name = entry.get().strip()
            if not new_name:
                messagebox.showwarning("Ошибка", "Имя не может быть пустым", parent=dialog)
                return
            if new_name != old_name and new_name in self.employees_list:
                messagebox.showwarning("Ошибка", "Такой сотрудник уже существует", parent=dialog)
                return
            self.employees_list[self.employees_list.index(old_name)] = new_name
            for item in self.inventory_data:
                if item.get('assignment') == old_name:
                    item['assignment'] = new_name
            for serial, records in self.history_data.items():
                for rec in records:
                    if rec.get('assignment') == old_name:
                        rec['assignment'] = new_name
            self.save_employees(self.employees_list)
            self.save_data()
            self.save_history()
            self.update_employee_comboboxes()
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="Сохранить", command=save_edit, style='Small.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Отмена", command=dialog.destroy, style='Small.TButton').pack(side='left', padx=5)

        dialog.wait_window()

    def delete_selected_employee(self):
        selection = self.all_employees_listbox.curselection()
        if not selection:
            return
        employee_name = self.all_employees_listbox.get(selection[0])
        if not employee_name:
            return
        in_use = any(item.get('assignment', '') == employee_name for item in self.inventory_data)
        if in_use:
            messagebox.showerror("Ошибка", f"Сотрудник '{employee_name}' используется в записях. Удаление запрещено.")
            return
        if messagebox.askyesno("Подтверждение", f"Удалить сотрудника '{employee_name}'?"):
            self.employees_list.remove(employee_name)
            self.save_employees(self.employees_list)
            self.unsaved_changes = True
            self.update_window_title()
            self.update_employee_comboboxes()

    def export_employees_to_excel(self):
        if not self.employees_list:
            messagebox.showwarning("Предупреждение", "Список сотрудников пуст")
            return
        try:
            import pandas as pd
            df = pd.DataFrame(sorted(self.employees_list), columns=["Сотрудник"])
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Сохранить список сотрудников",
                initialdir=self.data_dir
            )
            if not filename:
                return
            df.to_excel(filename, index=False, engine='openpyxl')
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            ws = wb.active
            ws.column_dimensions['A'].width = 30
            wb.save(filename)
            webbrowser.open(filename)
            messagebox.showinfo("Успех", f"Список сотрудников сохранён:\n{filename}")
        except ImportError:
            messagebox.showerror("Ошибка", "Установите openpyxl: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать:\n{e}")

    def filter_employees_by_search(self, event=None):
        search_term = self.employee_search_var.get().lower().strip()
        filtered_employees = [emp for emp in [""] + sorted(self.employees_list) if search_term in emp.lower()]
        self.employee_combo['values'] = filtered_employees
        if filtered_employees:
            self.employee_combo.current(0)
        else:
            self.employee_combo.set('')
        self.show_employee_equipment()

    def load_employees_from_inventory(self):
        if not self.inventory_data:
            messagebox.showwarning("Предупреждение", "Нет данных в базе инвентаризации.")
            return
        new_employees = set()
        for item in self.inventory_data:
            assignment = item.get('assignment', '').strip()
            if assignment:
                new_employees.add(assignment)
        if not new_employees:
            messagebox.showinfo("Информация", "В базе нет записей с закреплёнными сотрудниками.")
            return
        added_count = 0
        for emp in new_employees:
            if emp not in self.employees_list:
                self.employees_list.append(emp)
                added_count += 1
        if self.save_employees(self.employees_list):
            self.unsaved_changes = True
            self.update_window_title()
            self.update_employee_comboboxes()
            messagebox.showinfo("Успех",
                                f"Из базы загружено {added_count} новых сотрудников.\nВсего сотрудников: {len(self.employees_list)}")
        else:
            messagebox.showerror("Ошибка", "Не удалось сохранить список сотрудников.")

    def refresh_employee_list(self):
        self.update_employee_comboboxes()
        self.all_employees_listbox.delete(0, tk.END)
        for emp in sorted(self.employees_list):
            self.all_employees_listbox.insert(tk.END, emp)

    # =============== ФУНКЦИЯ ПЕРЕДАЧИ ОБОРУДОВАНИЯ ===============
    def transfer_selected_item(self):
        current_tab = self.notebook.index(self.notebook.select())
        tree = {0: self.all_tree, 2: self.search_tree, 3: self.employee_tree}.get(current_tab)
        if not tree:
            return
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("Предупреждение", "Выберите запись для передачи")
            return
        item_values = tree.item(selected_items[0], 'values')
        serial_number = item_values[2]
        target_item = next((item for item in self.inventory_data if item.get('serial_number') == serial_number), None)
        if not target_item:
            messagebox.showerror("Ошибка", "Запись не найдена в базе")
            return
        self.open_transfer_dialog(target_item)

    def open_transfer_dialog(self, equipment_item):
        dialog = tk.Toplevel(self.root)
        dialog.title("Передача оборудования")
        dialog.geometry("800x700")
        dialog.transient(self.root)
        dialog.grab_set()

        large_font = tkFont.Font(family='Arial', size=16)

        info_frame = ttk.LabelFrame(dialog, text="Информация об оборудовании", padding=15)
        info_frame.pack(fill='x', padx=20, pady=(20, 10))

        info_labels = [
            ("Тип:", equipment_item.get('equipment_type', '')),
            ("Модель:", equipment_item.get('model', '')),
            ("Серийный номер:", equipment_item.get('serial_number', '')),
            ("Текущее закрепление:", equipment_item.get('assignment', '')),
            ("Дата закрепления:", equipment_item.get('date', '')),
            ("Комментарии:", equipment_item.get('comments', ''))
        ]
        for i, (label_text, value) in enumerate(info_labels):
            ttk.Label(info_frame, text=label_text, font=large_font).grid(row=i, column=0, sticky='w', padx=(0, 20),
                                                                         pady=5)
            ttk.Label(info_frame, text=value, font=large_font, wraplength=500).grid(row=i, column=1, sticky='w', pady=5)

        transfer_frame = ttk.LabelFrame(dialog, text="Новое закрепление", padding=15)
        transfer_frame.pack(fill='x', padx=20, pady=10)

        ttk.Label(transfer_frame, text="Сотрудник:", font=large_font).pack(anchor='w', pady=(0, 10))
        new_employee_var = tk.StringVar()
        employee_combo = ttk.Combobox(transfer_frame, textvariable=new_employee_var,
                                      values=sorted(self.employees_list), state='readonly', font=large_font)
        employee_combo.pack(fill='x', pady=(0, 15))
        if self.employees_list:
            employee_combo.set(self.employees_list[0])

        date_var = tk.StringVar(value=datetime.now().strftime("%d.%m.%Y"))
        ttk.Label(transfer_frame, text="Дата передачи:", font=large_font).pack(anchor='w', pady=(0, 10))
        date_entry = ttk.Entry(transfer_frame, textvariable=date_var, font=large_font)
        date_entry.pack(fill='x', pady=(0, 15))

        def confirm_transfer():
            new_employee = new_employee_var.get().strip()
            transfer_date = date_var.get().strip()
            if not new_employee:
                messagebox.showwarning("Ошибка", "Выберите сотрудника", parent=dialog)
                return
            if not transfer_date:
                messagebox.showwarning("Ошибка", "Укажите дату передачи", parent=dialog)
                return
            try:
                datetime.strptime(transfer_date, "%d.%m.%Y")
            except ValueError:
                messagebox.showerror("Ошибка", "Неверный формат даты. Используйте дд.мм.гггг", parent=dialog)
                return

            old_assignment = equipment_item.get('assignment', '')
            equipment_item['assignment'] = new_employee
            equipment_item['date'] = transfer_date
            self.add_to_history(equipment_item['serial_number'], new_employee, transfer_date)
            self.save_data()
            self.save_history()
            self.show_all_data()
            self.refresh_employee_list()
            self.update_serial_combobox()
            self.update_history_combobox()
            messagebox.showinfo("Успех", f"Оборудование передано сотруднику {new_employee}", parent=dialog)
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="Подтвердить", command=confirm_transfer, style='Big.TButton').pack(side='left',
                                                                                                      padx=10)
        ttk.Button(btn_frame, text="Отмена", command=dialog.destroy, style='Big.TButton').pack(side='left', padx=10)

        dialog.wait_window()

    # =============== ОСТАЛЬНЫЕ МЕТОДЫ ===============
    def create_add_tab(self):
        fields = [
            ("Тип оборудования", "equipment_type"),
            ("Модель", "model"),
            ("Серийный номер", "serial_number"),
            ("Закрепление", "assignment"),
            ("Дата", "date"),
            ("Комментарии", "comments")
        ]
        self.entries = {}
        for i, (label_text, field_name) in enumerate(fields):
            label = ttk.Label(self.add_frame, text=label_text + ":")
            label.grid(row=i, column=0, sticky='w', padx=10, pady=5)
            if field_name == "equipment_type":
                self.equipment_type_var = tk.StringVar()
                combo = ttk.Combobox(self.add_frame, textvariable=self.equipment_type_var,
                                     values=sorted(self.equipment_types), width=38, font=self.default_font)
                combo.grid(row=i, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(combo)
                self.entries[field_name] = combo
            elif field_name == "comments":
                entry = scrolledtext.ScrolledText(self.add_frame, width=40, height=4, font=self.default_font)
                entry.grid(row=i, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(entry)
                self.entries[field_name] = entry
            elif field_name == "date":
                entry = ttk.Entry(self.add_frame, width=40, font=self.default_font)
                entry.insert(0, datetime.now().strftime("%d.%m.%Y"))
                entry.grid(row=i, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(entry)
                self.entries[field_name] = entry
            elif field_name == "assignment":
                self.assignment_var = tk.StringVar()
                self.assignment_combo = ttk.Combobox(self.add_frame, textvariable=self.assignment_var,
                                                     values=[""] + sorted(self.employees_list), width=38,
                                                     font=self.default_font)
                self.assignment_combo.grid(row=i, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(self.assignment_combo)
                self.entries[field_name] = self.assignment_combo
            else:
                entry = ttk.Entry(self.add_frame, width=40, font=self.default_font)
                entry.grid(row=i, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(entry)
                self.entries[field_name] = entry

        add_button = ttk.Button(self.add_frame, text="Добавить оборудование",
                                command=self.add_equipment, style='Big.TButton')
        add_button.grid(row=len(fields), column=0, columnspan=2, pady=20)

        self.add_frame.columnconfigure(1, weight=1)

    def open_add_employee_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Добавить сотрудника")
        dialog.geometry("500x200")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Введите имя сотрудника:", font=self.default_font).pack(pady=(20, 10))
        entry = ttk.Entry(dialog, width=45, font=self.default_font)
        entry.pack(pady=10)
        entry.focus()

        def on_add():
            name = entry.get().strip()
            if name:
                if self.add_employee(name):
                    self.update_employee_comboboxes()
                    dialog.destroy()
            else:
                messagebox.showwarning("Предупреждение", "Имя сотрудника не может быть пустым.", parent=dialog)

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="Добавить", command=on_add, style='Small.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Отмена", command=dialog.destroy, style='Small.TButton').pack(side='left', padx=5)

        dialog.wait_window()

    def create_equipment_tab(self):
        frame = self.equipment_frame
        ttk.Label(frame, text="Тип оборудования:", font=self.default_font).grid(row=0, column=0, sticky='w', padx=10,
                                                                                pady=5)
        self.equipment_type_entry = ttk.Entry(frame, width=40, font=self.default_font)
        self.equipment_type_entry.grid(row=0, column=1, padx=10, pady=5, sticky='we')
        self.bind_clipboard_events(self.equipment_type_entry)

        add_btn = ttk.Button(frame, text="➕ Добавить тип", command=self.add_equipment_type, style='Small.TButton')
        add_btn.grid(row=0, column=2, padx=10, pady=5)

        columns = ("Тип оборудования",)
        self.equipment_tree = ttk.Treeview(frame, columns=columns, show='headings', height=15)
        self.equipment_tree.heading("Тип оборудования", text="Тип оборудования")
        self.equipment_tree.column("Тип оборудования", width=300, anchor='center')
        self.equipment_tree.bind('<Button-3>', self.show_equipment_context_menu)

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.equipment_tree.yview)
        self.equipment_tree.configure(yscrollcommand=scrollbar.set)
        self.equipment_tree.grid(row=1, column=0, columnspan=3, padx=10, pady=5, sticky='nsew')
        scrollbar.grid(row=1, column=3, sticky='ns', pady=5)

        self.equipment_context_menu = tk.Menu(self.equipment_tree, tearoff=0)
        self.equipment_context_menu.add_command(label="🗑️ Удалить тип", command=self.delete_equipment_type)

        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(1, weight=1)

        self.refresh_equipment_list()

    def refresh_equipment_list(self):
        for item in self.equipment_tree.get_children():
            self.equipment_tree.delete(item)
        for eq_type in sorted(self.equipment_types):
            self.equipment_tree.insert("", "end", values=(eq_type,))

    def add_equipment_type(self):
        new_type = self.equipment_type_entry.get().strip()
        if not new_type:
            messagebox.showwarning("Предупреждение", "Введите название типа оборудования")
            return
        if new_type in self.equipment_types:
            messagebox.showwarning("Предупреждение", "Такой тип уже существует")
            return
        self.equipment_types.append(new_type)
        if self.save_equipment_types(self.equipment_types):
            messagebox.showinfo("Успех", "Тип оборудования добавлен")
            self.equipment_type_entry.delete(0, tk.END)
            self.refresh_equipment_list()
            self.unsaved_changes = True
            self.update_window_title()

    def show_equipment_context_menu(self, event):
        item = self.equipment_tree.identify_row(event.y)
        if item:
            self.equipment_tree.selection_set(item)
            self.equipment_context_menu.post(event.x_root, event.y_root)

    def delete_equipment_type(self):
        selected_items = self.equipment_tree.selection()
        if not selected_items:
            messagebox.showwarning("Предупреждение", "Выберите тип для удаления")
            return
        selected_type = self.equipment_tree.item(selected_items[0], 'values')[0]
        in_use = any(item.get('equipment_type') == selected_type for item in self.inventory_data)
        if in_use:
            messagebox.showerror("Ошибка", f"Тип '{selected_type}' используется в записях. Удаление запрещено.")
            return
        if messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить тип '{selected_type}'?"):
            self.equipment_types.remove(selected_type)
            if self.save_equipment_types(self.equipment_types):
                messagebox.showinfo("Успех", "Тип оборудования удален")
                self.refresh_equipment_list()
                self.unsaved_changes = True
                self.update_window_title()

    def create_settings_tab(self):
        frame = self.settings_frame
        ttk.Label(frame, text="Текущий каталог данных:", font=self.default_font).pack(pady=10)
        self.current_path_label = ttk.Label(frame, text=str(self.data_dir), font=self.default_font, wraplength=800)
        self.current_path_label.pack(pady=5)

        change_dir_btn = ttk.Button(frame, text="📁 Изменить каталог данных", command=self.choose_data_directory,
                                    style='Small.TButton')
        change_dir_btn.pack(pady=10)

        ttk.Label(frame, text="⚠️ Все файлы (inventory.json, history.json и др.) хранятся в этом каталоге",
                  font=self.default_font, foreground="red").pack(pady=10)

    def choose_data_directory(self):
        directory = filedialog.askdirectory(title="Выберите каталог для хранения данных")
        if not directory:
            return
        self.data_dir = Path(directory)
        self.save_settings(self.data_dir)
        self.current_path_label.config(text=str(self.data_dir))
        self.inventory_file = self.data_dir / "inventory.json"
        self.history_file = self.data_dir / "history.json"
        self.employees_file = self.data_dir / "sotrudniki.json"
        self.equipment_types_file = self.data_dir / "equipment_types.json"
        self.inventory_data = self.load_data()
        self.history_data = self.load_history()
        self.employees_list = self.load_employees()
        self.equipment_types = self.load_equipment_types()
        self.show_all_data()
        self.refresh_employee_list()
        self.update_history_combobox()
        self.update_serial_combobox()
        messagebox.showinfo("Успех", f"Каталог данных изменён:\n{self.data_dir}")

    def create_history_tab(self):
        frame = self.history_frame

        filter_frame = ttk.Frame(frame)
        filter_frame.pack(pady=10, fill='x', padx=10)

        ttk.Label(filter_frame, text="Сотрудник:", font=self.default_font).grid(row=0, column=0, sticky='w',
                                                                                padx=(0, 10))
        self.history_employee_var = tk.StringVar()
        self.history_employee_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.history_employee_var,
            values=[""] + sorted(
                set(item.get('assignment', '') for item in self.inventory_data if item.get('assignment'))),
            width=20,
            font=self.default_font
        )
        self.history_employee_combo.grid(row=0, column=1, padx=(0, 10), sticky='we')
        self.history_employee_combo.bind('<<ComboboxSelected>>', self.filter_history_by_employee)

        ttk.Label(filter_frame, text="Тип оборудования:", font=self.default_font).grid(row=0, column=2, sticky='w',
                                                                                       padx=(10, 10))
        self.history_type_var = tk.StringVar()
        self.history_type_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.history_type_var,
            values=[""] + sorted(
                set(item.get('equipment_type', '') for item in self.inventory_data if item.get('equipment_type'))),
            width=20,
            font=self.default_font
        )
        self.history_type_combo.grid(row=0, column=3, padx=(0, 10), sticky='we')
        self.history_type_combo.bind('<<ComboboxSelected>>', self.update_serial_combobox)

        ttk.Label(filter_frame, text="Серийный номер:", font=self.default_font).grid(row=0, column=4, sticky='w',
                                                                                     padx=(10, 10))
        self.history_serial_combo_var = tk.StringVar()
        self.history_serial_combo = ttk.Combobox(
            filter_frame,
            textvariable=self.history_serial_combo_var,
            values=[],
            width=20,
            font=self.default_font
        )
        self.history_serial_combo.grid(row=0, column=5, padx=(0, 10), sticky='we')
        self.history_serial_combo.bind('<<ComboboxSelected>>', self.show_history_for_equipment)

        refresh_btn = ttk.Button(filter_frame, text="🔄 Обновить", command=self.update_serial_combobox,
                                 style='Small.TButton')
        refresh_btn.grid(row=0, column=6, padx=(10, 0))

        for i in range(1, 6):
            filter_frame.columnconfigure(i, weight=1)

        history_label = ttk.Label(frame, text="История (по фильтру):", font=self.default_font)
        history_label.pack(anchor='w', padx=10, pady=(10, 0))

        # === ДОБАВЛЕН СТОЛБЕЦ "Модель" ===
        columns = ("Тип оборудования", "Модель", "Серийный номер", "Сотрудник", "Дата закрепления")
        self.history_tree = ttk.Treeview(frame, columns=columns, show='headings', height=10)
        # Настройка ширины колонок
        for col in columns:
            self.history_tree.heading(col, text=col)
            if col == "Сотрудник":
                width = 180
            elif col == "Модель":
                width = 200
            else:
                width = 150
            self.history_tree.column(col, width=width, anchor='center')

        scrollbar1 = ttk.Scrollbar(frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar1.set)
        self.history_tree.pack(side='top', fill='both', expand=True, padx=10, pady=5)
        scrollbar1.pack(side='right', fill='y', pady=5)

        export_filtered_frame = ttk.Frame(frame)
        export_filtered_frame.pack(pady=5, fill='x', padx=10)
        ttk.Button(export_filtered_frame, text="📥 Экспорт (по фильтру) в Excel",
                   command=self.export_filtered_history_to_excel, style='Small.TButton').pack(side='left', padx=5,
                                                                                              expand=True, fill='x')
        ttk.Button(export_filtered_frame, text="📄 Экспорт (по фильтру) в PDF",
                   command=self.export_filtered_history_to_pdf, style='Small.TButton').pack(side='left', padx=5,
                                                                                            expand=True, fill='x')

        all_history_label = ttk.Label(frame, text="Вся история (из history.json):", font=self.default_font)
        all_history_label.pack(anchor='w', padx=10, pady=(10, 0))

        # === ИСПРАВЛЕНИЕ: правильная упаковка Treeview и Scrollbar ===
        full_history_frame = ttk.Frame(frame)
        full_history_frame.pack(side='top', fill='both', expand=True, padx=10, pady=5)

        self.full_history_tree = ttk.Treeview(full_history_frame, columns=columns, show='headings', height=10)
        for col in columns:
            self.full_history_tree.heading(col, text=col)
            if col == "Сотрудник":
                width = 180
            elif col == "Модель":
                width = 200
            else:
                width = 150
            self.full_history_tree.column(col, width=width, anchor='center')

        scrollbar2 = ttk.Scrollbar(full_history_frame, orient="vertical", command=self.full_history_tree.yview)
        self.full_history_tree.configure(yscrollcommand=scrollbar2.set)
        self.full_history_tree.pack(side='left', fill='both', expand=True)
        scrollbar2.pack(side='right', fill='y')

        self.full_history_context_menu = tk.Menu(self.full_history_tree, tearoff=0)
        self.full_history_context_menu.add_command(label="🗑️ Удалить запись", command=self.delete_history_record)
        self.full_history_tree.bind('<Button-3>', self.show_full_history_context_menu)

        export_hist_btn = ttk.Button(frame, text="📥 Экспорт всей истории в Excel",
                                     command=self.export_history_to_excel, style='Small.TButton')
        export_hist_btn.pack(side='left', padx=5, pady=10)

        # === ДОБАВЛЕНА КНОПКА "НАЧАЛЬНОЕ ЗАПОЛНЕНИЕ" С ПОЯСНЕНИЕМ ===
        init_history_btn = ttk.Button(frame, text="🔄 Начальное заполнение",
                                      command=self.initialize_history_from_inventory, style='Small.TButton')
        init_history_btn.pack(side='left', padx=5, pady=10)

        hint_label = ttk.Label(frame, text="Нажимаем только при первом запуске программы!", foreground="gray", font=('Arial', 10))
        hint_label.pack(side='left', padx=5, pady=10)

        self.show_full_history()
        self.update_serial_combobox()

    def filter_history_by_employee(self, event=None):
        employee = self.history_employee_var.get().strip()
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        if not employee:
            return
        for serial, records in self.history_data.items():
            eq_type = "-"
            model = self._get_model_by_serial(serial)
            for inv_item in self.inventory_data:
                if inv_item.get('serial_number') == serial:
                    eq_type = inv_item.get('equipment_type', '-')
                    break
            for record in records:
                if record.get("assignment") == employee:
                    self.history_tree.insert("", "end", values=(
                        eq_type,
                        model,
                        serial,
                        record["assignment"],
                        record["date"]
                    ))

    def show_full_history(self):
        for item in self.full_history_tree.get_children():
            self.full_history_tree.delete(item)
        for serial, records in self.history_data.items():
            eq_type = "-"
            model = self._get_model_by_serial(serial)
            for inv_item in self.inventory_data:
                if inv_item.get('serial_number') == serial:
                    eq_type = inv_item.get('equipment_type', '-')
                    break
            for record in records:
                self.full_history_tree.insert("", "end", values=(
                    eq_type,
                    model,
                    serial,
                    record["assignment"],
                    record["date"]
                ))

    def show_full_history_context_menu(self, event):
        item = self.full_history_tree.identify_row(event.y)
        if item:
            self.full_history_tree.selection_set(item)
            self.full_history_context_menu.post(event.x_root, event.y_root)

    def delete_history_record(self):
        selected = self.full_history_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите запись для удаления")
            return
        if not messagebox.askyesno("Подтверждение", "Удалить выбранную запись из истории?"):
            return
        values = self.full_history_tree.item(selected[0], 'values')
        serial = values[2]
        assignment = values[3]
        date = values[4]
        if serial in self.history_data:
            self.history_data[serial] = [
                rec for rec in self.history_data[serial]
                if not (rec.get("assignment") == assignment and rec.get("date") == date)
            ]
            if not self.history_data[serial]:
                del self.history_data[serial]
            self.save_history()
            self.show_full_history()
            messagebox.showinfo("Успех", "Запись удалена из истории")

    def get_filtered_serial_numbers(self):
        selected_type = self.history_type_var.get()
        if not selected_type:
            serials = sorted(
                set(item.get('serial_number', '') for item in self.inventory_data if item.get('serial_number')))
        else:
            serials = sorted(set(item.get('serial_number', '') for item in self.inventory_data
                                 if item.get('equipment_type') == selected_type and item.get('serial_number')))
        return serials

    def update_serial_combobox(self, event=None):
        serials = self.get_filtered_serial_numbers()
        self.history_serial_combo['values'] = serials
        if serials:
            self.history_serial_combo_var.set(serials[0])
        else:
            self.history_serial_combo_var.set('')
        self.show_history_for_equipment()

    def show_history_for_equipment(self, event=None):
        serial = self.history_serial_combo_var.get().strip()
        if not serial:
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)
            return
        eq_type = "-"
        model = self._get_model_by_serial(serial)
        for item in self.inventory_data:
            if item.get('serial_number') == serial:
                eq_type = item.get('equipment_type', '-')
                break
        history_list = self.get_history_for_equipment(serial)
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        for record in history_list:
            self.history_tree.insert("", "end", values=(
                eq_type,
                model,
                serial,
                record["assignment"],
                record["date"]
            ))

    # =============== ЭКСПОРТ В EXCEL (общая функция) ===============
    def _export_to_excel(self, rows: List[Dict[str, Any]], filename_title: str, initial_dir: Path):
        try:
            import pandas as pd
            df = pd.DataFrame(rows)
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title=filename_title,
                initialdir=initial_dir
            )
            if not filename:
                return
            df.to_excel(filename, index=False, engine='openpyxl')
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            ws = wb.active
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            wb.save(filename)
            webbrowser.open(filename)
            messagebox.showinfo("Успех", f"Файл сохранён:\n{filename}")
        except ImportError:
            messagebox.showerror("Ошибка", "Установите openpyxl: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать:\n{e}")

    def export_history_to_excel(self):
        if not self.history_data:
            messagebox.showwarning("Предупреждение", "История пуста")
            return
        rows = []
        for serial, records in self.history_data.items():
            eq_type = "-"
            model = self._get_model_by_serial(serial)
            for item in self.inventory_data:
                if item.get('serial_number') == serial:
                    eq_type = item.get('equipment_type', '-')
                    break
            for record in records:
                rows.append({
                    "Тип оборудования": eq_type,
                    "Модель": model,
                    "Серийный номер": serial,
                    "Сотрудник": record["assignment"],
                    "Дата закрепления": record["date"]
                })
        self._export_to_excel(rows, "Сохранить историю в Excel", self.data_dir)

    def export_filtered_history_to_excel(self):
        items = self.history_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        rows = []
        for item in items:
            values = self.history_tree.item(item, 'values')
            rows.append({
                "Тип оборудования": values[0],
                "Модель": values[1],
                "Серийный номер": values[2],
                "Сотрудник": values[3],
                "Дата закрепления": values[4]
            })
        self._export_to_excel(rows, "Сохранить отфильтрованную историю в Excel", self.data_dir)

    def export_employees_to_excel(self):
        if not self.employees_list:
            messagebox.showwarning("Предупреждение", "Список сотрудников пуст")
            return
        rows = [{"Сотрудник": emp} for emp in sorted(self.employees_list)]
        self._export_to_excel(rows, "Сохранить список сотрудников", self.data_dir)

    def export_search_results_to_excel(self):
        items = self.search_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        rows = []
        for item in items:
            values = self.search_tree.item(item, 'values')
            rows.append({
                "Тип": values[0],
                "Модель": values[1],
                "Серийный номер": values[2],
                "Закрепление": values[3],
                "Дата": values[4],
                "Комментарии": values[5]
            })
        self._export_to_excel(rows, "Сохранить результаты поиска в Excel", self.data_dir)

    def export_employee_results_to_excel(self):
        items = self.employee_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        rows = []
        for item in items:
            values = self.employee_tree.item(item, 'values')
            rows.append({
                "Тип": values[0],
                "Модель": values[1],
                "Серийный номер": values[2],
                "Дата": values[3],
                "Комментарии": values[4]
            })
        self._export_to_excel(rows, "Сохранить отчёт по сотруднику в Excel", self.data_dir)

    def export_transfers_to_excel(self):
        items = self.transfers_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        rows = []
        for item in items:
            values = self.transfers_tree.item(item, 'values')
            rows.append({
                "Тип": values[0],
                "Серийный номер": values[1],
                "От кого": values[2],
                "Кому": values[3],
                "Дата передачи": values[4]
            })
        self._export_to_excel(rows, "Сохранить отчёт по передачам", self.data_dir)

    def export_to_excel(self):
        if not self.inventory_data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        active_tab = self.notebook.index(self.notebook.select())
        if active_tab == 0:
            current_data = []
            for item in self.all_tree.get_children():
                values = self.all_tree.item(item, 'values')
                current_data.append({
                    'Тип': values[0],
                    'Модель': values[1],
                    'Серийный номер': values[2],
                    'Закрепление': values[3],
                    'Дата': values[4],
                    'Комментарии': values[5]
                })
            data_to_export = current_data
        else:
            data_to_export = [
                {
                    'Тип': item.get('equipment_type', ''),
                    'Модель': item.get('model', ''),
                    'Серийный номер': item.get('serial_number', ''),
                    'Закрепление': item.get('assignment', ''),
                    'Дата': item.get('date', ''),
                    'Комментарии': item.get('comments', '')
                }
                for item in self.inventory_data
            ]
        self._export_to_excel(data_to_export, "Сохранить отчет в Excel", self.data_dir)

    # =============== ЭКСПОРТ В PDF (общая функция) ===============
    def _export_to_pdf(self, title: str, columns: List[str], data_rows: List[List[str]], subtitle: str = ""):
        try:
            font_path = _get_asset_path('ChakraPetch-Regular.ttf')
            pdf = PDFWithCyrillic(orientation='L')
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            pdf.set_font("ChakraPetch", '', 18)
            pdf.cell(0, 10, title, 0, 1, 'C')
            if subtitle:
                pdf.set_font("ChakraPetch", '', 12)
                pdf.cell(0, 10, subtitle, 0, 1, 'C')
            pdf.ln(10)
            pdf.set_font("ChakraPetch", '', 12)
            col_widths = []
            for col_index in range(len(columns)):
                max_width = pdf.get_string_width(columns[col_index]) + 6
                for row in data_rows:
                    w = pdf.get_string_width(str(row[col_index])) + 6
                    if w > max_width:
                        max_width = w
                col_widths.append(max_width)
            pdf.set_font("ChakraPetch", '', 14)
            for i, col in enumerate(columns):
                pdf.cell(col_widths[i], 10, col, 1, new_x="RIGHT", new_y="TOP", align='C')
            pdf.ln()
            pdf.set_font("ChakraPetch", '', 12)
            for row in data_rows:
                for i, cell_text in enumerate(row):
                    pdf.cell(col_widths[i], 10, str(cell_text), 1, new_x="RIGHT", new_y="TOP")
                pdf.ln()
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Сохранить отчёт в PDF",
                initialdir=self.data_dir
            )
            if not filename:
                return
            pdf.output(filename)
            webbrowser.open(filename)
            messagebox.showinfo("Успех", f"Отчёт сохранён:\n{filename}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать PDF отчёт: {e}")

    def export_to_pdf(self):
        if not self.inventory_data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        active_tab = self.notebook.index(self.notebook.select())
        if active_tab == 0:
            data_rows = []
            for item in self.all_tree.get_children():
                values = self.all_tree.item(item, 'values')
                row = [
                    values[0] or '-',
                    values[1] or '-',
                    values[2] or '-',
                    values[3] or '-',
                    values[4] or '-',
                    (values[5][:50] + '...') if values[5] and len(values[5]) > 50 else (values[5] or '-')
                ]
                data_rows.append(row)
        else:
            data_rows = []
            for item in self.inventory_data:
                row = [
                    item.get('equipment_type', '') or '-',
                    item.get('model', '') or '-',
                    item.get('serial_number', '') or '-',
                    item.get('assignment', '') or '-',
                    item.get('date', '') or '-',
                    (item.get('comments', '')[:50] + '...') if item.get('comments') and len(
                        item.get('comments')) > 50 else (item.get('comments', '') or '-')
                ]
                data_rows.append(row)
        total_equipment = len(self.inventory_data)
        unique_employees = len(
            set(item.get('assignment', '') for item in self.inventory_data if item.get('assignment')))
        subtitle = f"Всего единиц: {total_equipment} | Сотрудников: {unique_employees} | {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        self._export_to_pdf("Полный отчет по инвентаризации оборудования",
                            ["Тип", "Модель", "Серийный номер", "Закрепление", "Дата", "Комментарии"],
                            data_rows, subtitle)

    def export_search_results_to_pdf(self):
        items = self.search_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        data_rows = [list(self.search_tree.item(item, 'values')) for item in items]
        self._export_to_pdf("Отчет по результатам поиска оборудования",
                            ["Тип", "Модель", "Серийный номер", "Закрепление", "Дата", "Комментарии"],
                            data_rows)

    def export_employee_results_to_pdf(self):
        items = self.employee_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        employee_name = self.employee_var.get()
        data_rows = [list(self.employee_tree.item(item, 'values')) for item in items]
        self._export_to_pdf(f"Отчет по оборудованию сотрудника: {employee_name}",
                            ["Тип", "Модель", "Серийный номер", "Дата", "Комментарии"],
                            data_rows)

    def export_filtered_history_to_pdf(self):
        items = self.history_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        data_rows = [list(self.history_tree.item(item, 'values')) for item in items]
        self._export_to_pdf("Отчет по истории (по фильтру)",
                            ["Тип оборудования", "Модель", "Серийный номер", "Сотрудник", "Дата закрепления"],
                            data_rows)

    def export_transfers_to_pdf(self):
        items = self.transfers_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        start = self.transfers_start_var.get()
        end = self.transfers_end_var.get()
        subtitle = f"Период: с {start} по {end}"
        data_rows = [list(self.transfers_tree.item(item, 'values')) for item in items]
        self._export_to_pdf("Отчёт по передачам оборудования",
                            ["Тип", "Серийный номер", "От кого", "Кому", "Дата передачи"],
                            data_rows, subtitle)

    def create_about_tab(self):
        center_frame = ttk.Frame(self.about_frame)
        center_frame.pack(expand=True, fill='both')
        info_text = """Система инвентаризации оборудования — Inventory версия 2.0
        Разработано: Разин Григорий   Email: lantester35@gmail.com
        Программа предназначена для учёта и управления парком компьютерного оборудования:
        - Добавление, редактирование и удаление записей об оборудовании (тип, модель, серийный номер и т.д.)
        - Закрепление оборудования за сотрудниками
        - Ведение истории передач оборудования между сотрудниками
        - Экспорт данных в Excel и PDF
        - Генерация отчётов по передачам за заданный период
        - Визуализация статистики по типам оборудования
        - Автоматическое резервное копирование
        🔹 ВАЖНО ПРИ ПЕРВОМ ЗАПУСКЕ:
        1. На вкладке «Сотрудники» нажмите кнопку «📥 Из базы» — она загрузит всех сотрудников из существующих записей в инвентаризации.
        2. На вкладке «История» нажмите кнопку «🔄 Начальное заполнение» — она создаст историю первоначального закрепления оборудования на основе данных из inventory.json.
        Эти действия нужно выполнить один раз, чтобы система корректно заполнила справочники и историю. В дальнейшем они не требуются."""
        about_text = scrolledtext.ScrolledText(center_frame, width=100, height=17,
                                               font=self.default_font, wrap=tk.WORD)
        about_text.insert('1.0', info_text)
        about_text.config(state='disabled')
        about_text.pack(pady=20, padx=20, expand=True)

    def create_transfers_tab(self):
        frame = self.transfers_frame
        control_frame = ttk.LabelFrame(frame, text="Задайте период", padding=10)
        control_frame.pack(fill='x', padx=10, pady=10)

        ttk.Label(control_frame, text="Дата начала (дд.мм.гггг):", font=self.default_font).grid(row=0, column=0,
                                                                                                sticky='w',
                                                                                                padx=(0, 10))
        self.transfers_start_var = tk.StringVar(value=datetime.now().strftime("%d.%m.%Y"))
        start_entry = ttk.Entry(control_frame, textvariable=self.transfers_start_var, width=12, font=self.default_font)
        start_entry.grid(row=0, column=1, padx=(0, 20))

        ttk.Label(control_frame, text="Дата окончания (дд.мм.гггг):", font=self.default_font).grid(row=0, column=2,
                                                                                                   sticky='w',
                                                                                                   padx=(0, 10))
        self.transfers_end_var = tk.StringVar(value=datetime.now().strftime("%d.%m.%Y"))
        end_entry = ttk.Entry(control_frame, textvariable=self.transfers_end_var, width=12, font=self.default_font)
        end_entry.grid(row=0, column=3, padx=(0, 20))

        gen_btn = ttk.Button(control_frame, text="Сформировать отчёт", command=self.generate_transfers_report,
                             style='Small.TButton')
        gen_btn.grid(row=0, column=4, padx=(10, 0))

        result_frame = ttk.LabelFrame(frame, text="Результаты", padding=10)
        result_frame.pack(fill='both', expand=True, padx=10, pady=(0, 10))

        columns = ("Тип", "Серийный номер", "От кого", "Кому", "Дата передачи")
        self.transfers_tree = ttk.Treeview(result_frame, columns=columns, show='headings', height=15)
        for col in columns:
            self.transfers_tree.heading(col, text=col)
            self.transfers_tree.column(col, width=160, anchor='center')

        scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.transfers_tree.yview)
        self.transfers_tree.configure(yscrollcommand=scrollbar.set)
        self.transfers_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        export_frame = ttk.Frame(frame)
        export_frame.pack(fill='x', padx=10, pady=5)
        ttk.Button(export_frame, text="📥 Экспорт в Excel", command=self.export_transfers_to_excel,
                   style='Small.TButton').pack(side='left', padx=5)
        ttk.Button(export_frame, text="📄 Экспорт в PDF", command=self.export_transfers_to_pdf,
                   style='Small.TButton').pack(side='left', padx=5)

    def generate_transfers_report(self):
        try:
            start_str = self.transfers_start_var.get().strip()
            end_str = self.transfers_end_var.get().strip()
            start_dt = datetime.strptime(start_str, "%d.%m.%Y")
            end_dt = datetime.strptime(end_str, "%d.%m.%Y")
            if start_dt > end_dt:
                messagebox.showerror("Ошибка", "Дата начала не может быть позже даты окончания.")
                return
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты. Используйте дд.мм.гггг")
            return

        for item in self.transfers_tree.get_children():
            self.transfers_tree.delete(item)

        transfers = []
        for item in self.inventory_data:
            serial = item.get('serial_number')
            if not serial:
                continue
            # Собираем полную историю: начальное закрепление + history
            full_history = []
            # Добавляем первоначальное закрепление из inventory.json
            first_assignment = item.get('assignment')
            first_date = item.get('date')
            if first_assignment and first_date:
                try:
                    datetime.strptime(first_date, "%d.%m.%Y")
                    full_history.append({"assignment": first_assignment, "date": first_date})
                except ValueError:
                    pass  # некорректная дата — пропускаем
            # Добавляем записи из history.json
            history_records = self.history_data.get(serial, [])
            for rec in history_records:
                if rec.get("assignment") and rec.get("date"):
                    try:
                        datetime.strptime(rec["date"], "%d.%m.%Y")
                        full_history.append(rec)
                    except ValueError:
                        continue
            # Убираем дубликаты и сортируем по дате
            seen = set()
            unique_history = []
            for rec in full_history:
                key = (rec["assignment"], rec["date"])
                if key not in seen:
                    seen.add(key)
                    unique_history.append(rec)
            if len(unique_history) < 2:
                continue
            sorted_history = sorted(
                unique_history,
                key=lambda x: datetime.strptime(x["date"], "%d.%m.%Y")
            )
            # Формируем передачи: от предыдущего к текущему
            for i in range(1, len(sorted_history)):
                prev_emp = sorted_history[i - 1]["assignment"]
                curr_emp = sorted_history[i]["assignment"]
                transfer_date_str = sorted_history[i]["date"]
                if prev_emp == curr_emp:
                    continue
                try:
                    transfer_date = datetime.strptime(transfer_date_str, "%d.%m.%Y")
                except ValueError:
                    continue
                if start_dt <= transfer_date <= end_dt:
                    eq_type = item.get("equipment_type", "-")
                    transfers.append({
                        "equipment_type": eq_type,
                        "serial": serial,
                        "from": prev_emp,
                        "to": curr_emp,
                        "date": transfer_date_str
                    })

        if not transfers:
            messagebox.showinfo("Информация", "В указанный период передач оборудования не найдено.")
            return

        # Сортируем по дате передачи
        transfers.sort(key=lambda x: datetime.strptime(x["date"], "%d.%m.%Y"))
        for tr in transfers:
            self.transfers_tree.insert("", "end", values=(
                tr["equipment_type"],
                tr["serial"],
                tr["from"],
                tr["to"],
                tr["date"]
            ))

    def treeview_sort_column(self, tree, col, reverse):
        date_columns = ['Дата']
        def sort_key(val):
            if col in date_columns:
                try:
                    return datetime.strptime(val, "%d.%m.%Y")
                except:
                    return datetime.min
            else:
                return val.lower() if isinstance(val, str) else val
        data = [(tree.set(k, col), k) for k in tree.get_children('')]
        data.sort(key=lambda t: sort_key(t[0]), reverse=reverse)
        for index, (_, k) in enumerate(data):
            tree.move(k, '', index)
        tree.heading(col, command=lambda: self.treeview_sort_column(tree, col, not reverse))

    def show_context_menu(self, event):
        tree = event.widget
        item = tree.identify_row(event.y)
        if item:
            tree.selection_set(item)
            if tree == self.search_tree:
                self.search_context_menu.post(event.x_root, event.y_root)
            elif tree == self.employee_tree:
                self.employee_context_menu.post(event.x_root, event.y_root)
            elif tree == self.all_tree:
                self.all_context_menu.post(event.x_root, event.y_root)

    def delete_selected_item(self, event=None):
        current_tab = self.notebook.index(self.notebook.select())
        tree = {0: self.all_tree, 2: self.search_tree, 3: self.employee_tree}.get(current_tab)
        if not tree:
            return
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("Предупреждение", "Выберите запись для удаления")
            return
        if not messagebox.askyesno("Подтверждение", "Вы уверены, что хотите удалить выбранную запись?"):
            return
        for item in selected_items:
            values = tree.item(item, 'values')
            serial_number = values[2] if len(values) > 2 else None
            if not serial_number:
                continue
            for i, inv_item in enumerate(self.inventory_data):
                if inv_item.get('serial_number') == serial_number:
                    del self.inventory_data[i]
                    tree.delete(item)
                    break
        if self.save_data():
            messagebox.showinfo("Успех", "Запись успешно удалена")
            self.refresh_employee_list()
            self.show_all_data()
            self.update_history_combobox()
            self.update_serial_combobox()

    def add_equipment(self):
        equipment_data = {}
        for field_name, entry in self.entries.items():
            if field_name == "comments":
                equipment_data[field_name] = entry.get("1.0", tk.END).strip()
            else:
                equipment_data[field_name] = entry.get().strip()

        if not equipment_data.get('equipment_type') or not equipment_data.get('serial_number'):
            messagebox.showwarning("Предупреждение", "Заполните обязательные поля: Тип оборудования и Серийный номер")
            return

        if equipment_data['equipment_type'] not in self.equipment_types:
            messagebox.showwarning("Предупреждение", "Выбранный тип оборудования не существует в списке.")
            return

        if not self.is_serial_number_unique(equipment_data['serial_number']):
            messagebox.showerror("Ошибка", f"Серийный номер '{equipment_data['serial_number']}' уже существует!")
            return

        equipment_data['created_datetime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.inventory_data.append(equipment_data)
        self.add_to_history(equipment_data['serial_number'], equipment_data['assignment'], equipment_data['date'])
        if self.save_data():
            messagebox.showinfo("Успех", "Оборудование успешно добавлено!")
            self.clear_entries()
            self.refresh_employee_list()
            self.show_all_data()
            self.update_history_combobox()
            self.update_serial_combobox()

    def clear_entries(self):
        for field_name, entry in self.entries.items():
            if field_name == "comments":
                entry.delete("1.0", tk.END)
            elif field_name == "date":
                entry.delete(0, tk.END)
                entry.insert(0, datetime.now().strftime("%d.%m.%Y"))
            else:
                entry.delete(0, tk.END)
        if 'equipment_type' in self.entries:
            self.entries['equipment_type'].set('')

    def perform_search(self, event=None):
        search_text = self.search_entry.get().lower().strip()
        selected_employee = self.search_employee_var.get().strip()
        for item in self.search_tree.get_children():
            self.search_tree.delete(item)
        if not search_text and not selected_employee:
            return
        for item in self.inventory_data:
            matches_search = any(search_text in str(value).lower() for value in item.values() if value)
            matches_employee = (not selected_employee) or (item.get('assignment', '') == selected_employee)
            if matches_search and matches_employee:
                self.search_tree.insert("", "end", values=(
                    item.get('equipment_type', ''),
                    item.get('model', ''),
                    item.get('serial_number', ''),
                    item.get('assignment', ''),
                    item.get('date', ''),
                    (item.get('comments', '')[:50] + '...') if item.get('comments') and len(
                        item.get('comments')) > 50 else item.get('comments', '')
                ))

    def clear_search(self):
        self.search_entry.delete(0, tk.END)
        self.search_employee_var.set('')
        for item in self.search_tree.get_children():
            self.search_tree.delete(item)

    def update_history_combobox(self):
        serials = sorted(
            set(item.get('serial_number', '') for item in self.inventory_data if item.get('serial_number')))
        self.history_serial_combo['values'] = serials
        if serials:
            self.history_serial_combo_var.set(serials[0])
        else:
            self.history_serial_combo_var.set('')

    def show_employee_equipment(self, event=None):
        employee = self.employee_var.get()
        for item in self.employee_tree.get_children():
            self.employee_tree.delete(item)
        if not employee:
            return
        for item in self.inventory_data:
            if item.get('assignment') == employee:
                self.employee_tree.insert("", "end", values=(
                    item.get('equipment_type', ''),
                    item.get('model', ''),
                    item.get('serial_number', ''),
                    item.get('date', ''),
                    (item.get('comments', '')[:50] + '...') if item.get('comments') and len(
                        item.get('comments')) > 50 else item.get('comments', '')
                ))

    def show_all_data(self):
        for item in self.all_tree.get_children():
            self.all_tree.delete(item)
        self.inventory_data = self.load_data()
        for item in self.inventory_data:
            self.all_tree.insert("", "end", values=(
                item.get('equipment_type', ''),
                item.get('model', ''),
                item.get('serial_number', ''),
                item.get('assignment', ''),
                item.get('date', ''),
                (item.get('comments', '')[:50] + '...') if item.get('comments') and len(
                    item.get('comments')) > 50 else item.get('comments', '')
            ))
        self.refresh_employee_list()
        self.update_history_combobox()
        self.update_serial_combobox()
        self.treeview_sort_column(self.all_tree, "Закрепление", False)

    def on_tree_double_click(self, event):
        tree = event.widget
        item = tree.identify('item', event.x, event.y)
        column = tree.identify_column(event.x)
        if not item or item not in tree.get_children():
            return
        col_index = int(column.replace('#', '')) - 1
        current_values = tree.item(item, 'values')
        current_value = current_values[col_index]
        field_names_map = {
            self.employee_tree: {0: 'equipment_type', 1: 'model', 2: 'serial_number', 3: 'date', 4: 'comments'},
            self.all_tree: {0: 'equipment_type', 1: 'model', 2: 'serial_number', 3: 'assignment', 4: 'date',
                            5: 'comments'},
            self.search_tree: {0: 'equipment_type', 1: 'model', 2: 'serial_number', 3: 'assignment', 4: 'date',
                               5: 'comments'}
        }
        field_names = field_names_map.get(tree)
        if not field_names:
            return
        field_name = field_names.get(col_index)
        if not field_name:
            return
        idx = next(
            (i for i, inv_item in enumerate(self.inventory_data) if inv_item.get('serial_number') == current_values[2]),
            None)
        if idx is None:
            return
        self.edit_cell(tree, item, col_index, field_name, current_value, idx)

    def edit_cell(self, tree, item, col_index, field_name, current_value, data_index):
        bbox = tree.bbox(item, column=f'#{col_index + 1}')
        if not bbox:
            return

        def validate_and_save(new_value: str):
            if field_name == 'date':
                try:
                    datetime.strptime(new_value, "%d.%m.%Y")
                except ValueError:
                    messagebox.showerror("Ошибка", "Неверный формат даты. Используйте дд.мм.гггг")
                    return False
            elif field_name == 'serial_number':
                if not self.is_serial_number_unique(new_value, exclude_index=data_index):
                    messagebox.showerror("Ошибка", f"Серийный номер '{new_value}' уже существует!")
                    return False
            return True

        if field_name == 'comments':
            text_edit = scrolledtext.ScrolledText(tree, width=40, height=4, font=self.default_font)
            text_edit.insert('1.0', current_value)
            text_edit.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3] * 3)
            text_edit.focus()
            self.bind_clipboard_events(text_edit)

            def save_edit(event=None):
                new_value = text_edit.get('1.0', tk.END).strip()
                if validate_and_save(new_value):
                    text_edit.destroy()
                    current_values = list(tree.item(item, 'values'))
                    current_values[col_index] = new_value
                    tree.item(item, values=current_values)
                    self.inventory_data[data_index][field_name] = new_value
                    self.unsaved_changes = True
                    self.update_window_title()
                    self.save_data()

            def cancel_edit(event=None):
                text_edit.destroy()

            text_edit.bind('<Return>', save_edit)
            text_edit.bind('<Escape>', cancel_edit)
            text_edit.bind('<FocusOut>', lambda e: save_edit())

        elif field_name == 'assignment':
            combo_edit = ttk.Combobox(
                tree,
                values=[""] + sorted(self.employees_list),
                width=bbox[2] // 8,
                font=self.default_font,
                state='readonly'
            )
            combo_edit.set(current_value)
            combo_edit.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            combo_edit.focus()

            def save_edit(event=None):
                new_value = combo_edit.get().strip()
                if validate_and_save(new_value):
                    combo_edit.destroy()
                    current_values = list(tree.item(item, 'values'))
                    current_values[col_index] = new_value
                    tree.item(item, values=current_values)
                    old_value = current_value
                    self.inventory_data[data_index][field_name] = new_value
                    serial_number = self.inventory_data[data_index].get('serial_number', '')
                    if serial_number and old_value != new_value:
                        current_date = datetime.now().strftime("%d.%m.%Y")
                        self.add_to_history(serial_number, new_value, current_date)
                    self.unsaved_changes = True
                    self.update_window_title()
                    self.save_data()

            def cancel_edit(event=None):
                combo_edit.destroy()

            combo_edit.bind('<Return>', save_edit)
            combo_edit.bind('<Escape>', cancel_edit)
            combo_edit.bind('<FocusOut>', lambda e: save_edit())
            combo_edit.bind('<<ComboboxSelected>>', lambda e: save_edit())

        else:
            entry_edit = ttk.Entry(tree, width=bbox[2] // 8, font=self.default_font)
            entry_edit.insert(0, current_value)
            entry_edit.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            entry_edit.focus()
            self.bind_clipboard_events(entry_edit)

            def save_edit(event=None):
                new_value = entry_edit.get().strip()
                if validate_and_save(new_value):
                    entry_edit.destroy()
                    current_values = list(tree.item(item, 'values'))
                    current_values[col_index] = new_value
                    tree.item(item, values=current_values)
                    self.inventory_data[data_index][field_name] = new_value
                    self.unsaved_changes = True
                    self.update_window_title()
                    self.save_data()

            def cancel_edit(event=None):
                entry_edit.destroy()

            entry_edit.bind('<Return>', save_edit)
            entry_edit.bind('<Escape>', cancel_edit)
            entry_edit.bind('<FocusOut>', lambda e: save_edit())

    def show_equipment_graph(self):
        if not self.inventory_data:
            messagebox.showwarning("Предупреждение", "Нет данных для построения графика")
            return
        try:
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            type_counts = {}
            for item in self.inventory_data:
                eq_type = item.get('equipment_type', 'Не указано')
                type_counts[eq_type] = type_counts.get(eq_type, 0) + 1
            sorted_types = sorted(type_counts.items(), key=lambda x: x[1], reverse=True)
            types, counts = zip(*sorted_types) if sorted_types else ([], [])
            graph_window = tk.Toplevel(self.root)
            graph_window.title("График распределения оборудования")
            graph_window.geometry("900x700")
            fig, ax = plt.subplots(figsize=(10, 8))
            bars = ax.bar(types, counts, color='skyblue', edgecolor='black')
            ax.set_title('Количество единиц оборудования по типам', fontsize=16, fontweight='bold')
            ax.set_xlabel('Тип оборудования', fontsize=12)
            ax.set_ylabel('Количество', fontsize=12)
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2., height + 0.1,
                        f'{int(height)}', ha='center', va='bottom', fontsize=10)
            canvas = FigureCanvasTkAgg(fig, master=graph_window)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            close_btn = ttk.Button(graph_window, text="Закрыть", command=graph_window.destroy, style='Big.TButton')
            close_btn.pack(pady=10)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось построить график: {e}")

    def schedule_auto_save(self):
        def auto_save():
            if self.save_data():
                logger.info("[AUTO-SAVE] Данные сохранены")
            self.root.after(self.auto_save_interval, auto_save)
        self.root.after(self.auto_save_interval, auto_save)

def main():
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()