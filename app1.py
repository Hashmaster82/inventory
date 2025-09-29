import json
import os
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
from datetime import datetime
from fpdf import FPDF
import webbrowser
import tkinter.font as tkFont
import sys

# ----------------- Класс PDF с поддержкой кириллицы -----------------
class PDFWithCyrillic(FPDF):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Получаем путь к шрифту относительно скрипта
        font_path = self._get_asset_path('ChakraPetch-Regular.ttf')
        if not os.path.exists(font_path):
            font_path = self._get_asset_path('assets/fonts/ChakraPetch-Regular.ttf')
            if not os.path.exists(font_path):
                raise FileNotFoundError(f"Шрифт не найден ни в корне проекта, ни в assets/fonts/: {font_path}")
        self.add_font('ChakraPetch', '', font_path, uni=True)

    def _get_asset_path(self, filename):
        """Получает путь к файлу внутри пакета или в рабочей директории.
           Ищет в порядке: корень проекта → assets/fonts/ → MEIPASS"""
        if getattr(sys, 'frozen', False):
            # Запущено как .exe
            base_path = sys._MEIPASS
            path1 = os.path.join(base_path, 'assets', 'fonts', filename)
            path2 = os.path.join(base_path, filename)
            if os.path.exists(path1):
                return path1
            elif os.path.exists(path2):
                return path2
            else:
                raise FileNotFoundError(f"Шрифт не найден в MEIPASS: {filename}")
        else:
            # Запущено как скрипт .py
            base_path = os.path.dirname(__file__)
            path1 = os.path.join(base_path, filename)  # Корень проекта
            path2 = os.path.join(base_path, 'assets', 'fonts', filename)  # Подпапка
            if os.path.exists(path1):
                return path1
            elif os.path.exists(path2):
                return path2
            else:
                raise FileNotFoundError(f"Шрифт не найден ни в корне проекта, ни в assets/fonts/: {filename}")


# ----------------- Основной класс приложения -----------------
class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Система инвентаризации оборудования")
        self.root.state('zoomed')
        self.default_font = tkFont.Font(family='Arial', size=14)
        self.root.option_add("*Font", self.default_font)
        # Пути к файлам
        self.inventory_file = r"\\fs\SHARE_BH\it\inventory\inventory.json"
        self.equipment_types_file = "equipment_types.json"
        self.history_file = "history.json"
        # Убедимся, что каталог для inventory.json существует
        inventory_dir = os.path.dirname(self.inventory_file)
        os.makedirs(inventory_dir, exist_ok=True)
        # Автоматически создаем history.json в том же каталоге, что и inventory.json, если его нет
        history_dir = os.path.dirname(self.inventory_file)
        self.history_file = os.path.join(history_dir, "history.json")
        if not os.path.exists(self.history_file):
            try:
                with open(self.history_file, 'w', encoding='utf-8') as f:
                    json.dump({}, f)
                print(f"[INFO] Создан новый файл истории: {self.history_file}")
            except Exception as e:
                print(f"[ERROR] Не удалось создать history.json: {e}")
        # Загрузка данных
        self.inventory_data = self.load_data()
        self.equipment_types = self.load_equipment_types()
        self.history_data = self.load_history()
        # --- НОВЫЙ КОД: ЗАГРУЗКА СПИСКА СОТРУДНИКОВ ---
        self.employees_list = self.load_employees()  # Загружаем список сотрудников из sotrudniki.json
        # Создание интерфейса
        self.create_widgets()
        # --- Автосохранение каждые 5 минут ---
        self.auto_save_interval = 300000  # 5 минут в миллисекундах
        self.schedule_auto_save()

    # =============== РАБОТА С ТИПАМИ ОБОРУДОВАНИЯ ===============
    def load_equipment_types(self):
        try:
            if os.path.exists(self.equipment_types_file):
                with open(self.equipment_types_file, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    return data if isinstance(data, list) else []
            else:
                default_types = ["Монитор", "Сисблок", "МФУ", "Клавиатура", "Мышь", "Наушники"]
                self.save_equipment_types(default_types)
                return default_types
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить типы оборудования: {e}")
            print(f"[ERROR] Load equipment types: {e}")  # Лог в консоль
            return []

    def save_equipment_types(self, types_list):
        try:
            with open(self.equipment_types_file, 'w', encoding='utf-8') as file:
                json.dump(types_list, file, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить типы оборудования: {e}")
            print(f"[ERROR] Save equipment types: {e}")
            return False

    # =============== РАБОТА СО СПИСКОМ СОТРУДНИКОВ ===============
    def load_employees(self):
        """Загружает список сотрудников из файла sotrudniki.json"""
        try:
            employees_file = os.path.join(os.path.dirname(self.inventory_file), "sotrudniki.json")
            if os.path.exists(employees_file):
                with open(employees_file, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    return data if isinstance(data, list) else []
            else:
                # Создаем пустой файл, если его нет
                self.save_employees([])
                return []
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить список сотрудников: {e}")
            print(f"[ERROR] Load employees: {e}")
            return []

    def save_employees(self, employees_list):
        """Сохраняет список сотрудников в файл sotrudniki.json"""
        try:
            employees_file = os.path.join(os.path.dirname(self.inventory_file), "sotrudniki.json")
            with open(employees_file, 'w', encoding='utf-8') as file:
                json.dump(employees_list, file, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить список сотрудников: {e}")
            print(f"[ERROR] Save employees: {e}")
            return False

    def add_employee(self, employee_name):
        """Добавляет нового сотрудника в список и сохраняет"""
        if not employee_name.strip():
            return False
        employee_name = employee_name.strip()
        if employee_name in self.employees_list:
            messagebox.showwarning("Предупреждение", "Этот сотрудник уже существует.")
            return False
        self.employees_list.append(employee_name)
        self.save_employees(self.employees_list)
        return True

    def delete_employee(self, employee_name):
        """Удаляет сотрудника из списка (если он не используется)"""
        if not employee_name:
            return False
        # Проверяем, используется ли сотрудник в inventory_data
        in_use = any(item.get('assignment', '') == employee_name for item in self.inventory_data)
        if in_use:
            messagebox.showerror("Ошибка",
                                 f"Сотрудник '{employee_name}' используется в {in_use} записях. Удаление запрещено.")
            return False
        if messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить сотрудника '{employee_name}'?"):
            self.employees_list.remove(employee_name)
            self.save_employees(self.employees_list)
            return True
        return False

    def update_employee_comboboxes(self):
        """Обновляет все Combobox'ы, использующие список сотрудников"""
        # Обновляем Combobox в вкладке "Добавить оборудование"
        if hasattr(self, 'assignment_combo'):
            self.assignment_combo['values'] = [""] + sorted(self.employees_list)
            if self.employees_list:
                self.assignment_var.set(self.employees_list[0])  # Установить первый как значение по умолчанию
            else:
                self.assignment_var.set('')
        # Обновляем Combobox в вкладке "Сотрудники"
        if hasattr(self, 'employee_combo'):
            self.employee_combo['values'] = [""] + sorted(self.employees_list)
            if self.employees_list:
                self.employee_var.set(self.employees_list[0])
            else:
                self.employee_var.set('')
        # Обновляем комбобоксы в вкладке "История", чтобы они перечитали данные из inventory_data
        self.update_serial_combobox()

    # =============== РАБОТА С ИСТОРИЕЙ ===============
    def load_history(self):
        """Загружает историю закреплений из файла history.json"""
        try:
            if os.path.exists(self.history_file):
                with open(self.history_file, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    return data if isinstance(data, dict) else {}
            else:
                # Если файл отсутствует - создаём пустой словарь (это уже сделано при инициализации)
                return {}
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить историю: {e}")
            print(f"[ERROR] Load history: {e}")
            return {}

    def save_history(self):
        """Атомарно сохраняет историю в файл history.json"""
        try:
            temp_file = self.history_file + ".tmp"
            with open(temp_file, 'w', encoding='utf-8') as file:
                json.dump(self.history_data, file, ensure_ascii=False, indent=2)
            # Атомарная замена файла
            os.replace(temp_file, self.history_file)
            return True
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить историю: {e}")
            print(f"[ERROR] Save history: {e}")
            return False

    def add_to_history(self, serial_number, assignment, date):
        """Добавляет запись в историю закрепления оборудования"""
        if not serial_number or not assignment:
            return
        if serial_number not in self.history_data:
            self.history_data[serial_number] = []
        entry = {"assignment": assignment, "date": date}
        if entry not in self.history_data[serial_number]:
            self.history_data[serial_number].append(entry)
            self.save_history()

    def get_history_for_equipment(self, serial_number):
        """Возвращает историю закреплений для конкретного серийного номера"""
        return self.history_data.get(serial_number, [])

    # =============== ОСНОВНАЯ ЛОГИКА ===============
    def load_data(self):
        try:
            if os.path.exists(self.inventory_file):
                with open(self.inventory_file, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    if not isinstance(data, list):
                        raise ValueError("Файл должен содержать массив объектов")
                    return data
            else:
                os.makedirs(os.path.dirname(self.inventory_file), exist_ok=True)
                with open(self.inventory_file, 'w', encoding='utf-8') as file:
                    json.dump([], file)
                return []
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить данные: {e}")
            print(f"[ERROR] Load inventory data: {e}")
            return []

    def save_data(self):
        """Сохраняет данные в файл inventory.json"""
        try:
            os.makedirs(os.path.dirname(self.inventory_file), exist_ok=True)
            with open(self.inventory_file, 'w', encoding='utf-8') as file:
                json.dump(self.inventory_data, file, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить данные: {e}")
            print(f"[ERROR] Save inventory data: {e}")
            return False

    def create_backup(self):
        """Создание резервной копии JSON файла"""
        try:
            if not os.path.exists(self.inventory_file):
                messagebox.showwarning("Предупреждение", "Файл инвентаризации не существует")
                return
            backup_dir = os.path.join(os.path.dirname(self.inventory_file), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"inventory_backup_{timestamp}.json"
            backup_path = os.path.join(backup_dir, backup_filename)
            import shutil
            shutil.copy2(self.inventory_file, backup_path)
            messagebox.showinfo("Успех", f"Резервная копия создана:\n{backup_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать резервную копию: {e}")
            print(f"[ERROR] Create backup: {e}")

    def bind_clipboard_events(self, widget):
        def do_copy(event):
            widget.event_generate("<<Copy>>")
            return "break"

        def do_cut(event):
            widget.event_generate("<<Cut>>")
            return "break"

        def do_paste(event):
            widget.event_generate("<<Paste>>")
            return "break"

        widget.bind("<Control-c>", do_copy)
        widget.bind("<Control-x>", do_cut)
        widget.bind("<Control-v>", do_paste)

    def create_widgets(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        # ====== ВЕРХНЯЯ ПАНЕЛЬ КНОПОК ======
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10, fill='x')
        # Уменьшенные кнопки для основных действий
        button_style = ttk.Style()
        button_style.configure('Small.TButton', font=('Arial', 12), padding=6)
        backup_button = ttk.Button(button_frame, text="📂 Создать резервную копию",
                                   command=self.create_backup, style='Small.TButton')
        backup_button.pack(side='left', padx=5, fill='x', expand=True)
        pdf_button = ttk.Button(button_frame, text="📄 Выгрузить полный отчет в PDF",
                                command=self.export_to_pdf, style='Small.TButton')
        pdf_button.pack(side='left', padx=5, fill='x', expand=True)
        excel_button = ttk.Button(button_frame, text="📊 Выгрузить полный отчёт в Excel",
                                  command=self.export_to_excel, style='Small.TButton')
        excel_button.pack(side='left', padx=5, fill='x', expand=True)
        graph_button = ttk.Button(button_frame, text="📈 График",
                                  command=self.show_equipment_graph, style='Small.TButton')
        graph_button.pack(side='left', padx=5, fill='x', expand=True)
        # ====== НАБОР ВКЛАДОК ======
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True)
        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 18, 'bold'))
        style.configure('TNotebook.Tab', font=('Arial', 16, 'bold'), padding=[20, 10])
        # Создаём все фреймы вкладок
        self.show_all_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.show_all_frame, text="Показать всё")
        self.add_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.add_frame, text="Добавить оборудование")
        self.search_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.search_frame, text="Поиск оборудования")
        self.employee_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.employee_frame, text="Сотрудники")
        self.equipment_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.equipment_frame, text="Оборудование")
        self.settings_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_frame, text="Настройки")
        self.history_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.history_frame, text="История")
        self.about_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.about_frame, text="Инфо")
        # Создаём содержимое вкладок (в порядке зависимости)
        self.create_history_tab()  # Нужен для update_history_combobox()
        self.create_employee_tab()  # Может использовать update_history_combobox()
        self.create_show_all_tab()  # Загружает данные и обновляет другие комбобоксы
        self.create_add_tab()
        self.create_search_tab()
        self.create_equipment_tab()
        self.create_settings_tab()
        self.create_about_tab()
        # Устанавливаем первую вкладку
        self.notebook.select(self.show_all_frame)

    # =============== ВКЛАДКА: ПОКАЗАТЬ ВСЁ ===============
    def create_show_all_tab(self):
        # Уменьшенная кнопка "Обновить данные"
        refresh_button = ttk.Button(self.show_all_frame, text="Обновить данные",
                                    command=self.show_all_data, style='Small.TButton')
        refresh_button.pack(pady=5)
        table_frame = ttk.Frame(self.show_all_frame)
        table_frame.pack(fill='both', expand=True, padx=10, pady=5)
        columns = ("Тип", "Модель", "Серийный номер", "Закрепление", "Дата", "Комментарии")
        self.all_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=20)
        for col in columns:
            self.all_tree.heading(col, text=col,
                                  command=lambda _col=col: self.treeview_sort_column(self.all_tree, _col, False))
            self.all_tree.column(col, width=150, anchor='center')
        self.all_tree.bind('<Double-1>', self.on_tree_double_click)
        self.all_tree.bind('<Button-3>', self.show_context_menu)
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.all_tree.yview)
        self.all_tree.configure(yscrollcommand=scrollbar.set)
        self.all_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        self.all_context_menu = tk.Menu(self.all_tree, tearoff=0)
        self.all_context_menu.add_command(label="Удалить запись", command=self.delete_selected_item)
        self.show_all_data()  # Первоначальная загрузка

    # =============== ВКЛАДКА: ДОБАВИТЬ ОБОРУДОВАНИЕ ===============
    def create_add_tab(self):
        fields = [
            ("Тип оборудования", "equipment_type"),
            ("Модель", "model"),
            ("Серийный номер", "serial_number"),
            ("Закрепление", "assignment"),
            ("Дата", "date"),
            ("Комментарии", "comments")
        ]
        self.entries = {}
        row_offset = 0
        for i, (label_text, field_name) in enumerate(fields):
            label = ttk.Label(self.add_frame, text=label_text + ":")
            label.grid(row=i + row_offset, column=0, sticky='w', padx=10, pady=5)
            if field_name == "equipment_type":
                self.equipment_type_var = tk.StringVar()
                combo = ttk.Combobox(self.add_frame, textvariable=self.equipment_type_var,
                                     values=sorted(self.equipment_types), width=38, font=self.default_font)
                combo.grid(row=i + row_offset, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(combo)
                self.entries[field_name] = combo
            elif field_name == "comments":
                entry = scrolledtext.ScrolledText(self.add_frame, width=40, height=4, font=self.default_font)
                entry.grid(row=i + row_offset, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(entry)
                self.entries[field_name] = entry
            elif field_name == "date":
                entry = ttk.Entry(self.add_frame, width=40, font=self.default_font)
                entry.insert(0, datetime.now().strftime("%d.%m.%Y"))
                entry.grid(row=i + row_offset, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(entry)
                self.entries[field_name] = entry
            elif field_name == "assignment":
                self.assignment_var = tk.StringVar()
                self.assignment_combo = ttk.Combobox(self.add_frame, textvariable=self.assignment_var,
                                                     values=[""] + sorted(self.employees_list), width=38,
                                                     font=self.default_font)
                self.assignment_combo.grid(row=i + row_offset, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(self.assignment_combo)
                self.entries[field_name] = self.assignment_combo
            else:  # model, serial_number
                entry = ttk.Entry(self.add_frame, width=40, font=self.default_font)
                entry.grid(row=i + row_offset, column=1, padx=10, pady=5, sticky='we')
                self.bind_clipboard_events(entry)
                self.entries[field_name] = entry
        add_button = ttk.Button(self.add_frame, text="Добавить оборудование",
                                command=self.add_equipment, style='Big.TButton')
        add_button.grid(row=len(fields) + row_offset, column=0, columnspan=2, pady=20)
        self.add_frame.columnconfigure(1, weight=1)
        self.add_frame.rowconfigure(len(fields) + row_offset, weight=1)

    def open_add_employee_dialog(self):
        """Открывает диалог для добавления нового сотрудника"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Добавить сотрудника")
        dialog.geometry("500x200")  # Увеличено до 500x200
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()  # Модальный диалог
        # Центрирование диалога на экране
        dialog.update_idletasks()
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        window_width = dialog.winfo_width()
        window_height = dialog.winfo_height()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        dialog.geometry(f"+{x}+{y}")
        ttk.Label(dialog, text="Введите имя сотрудника:", font=self.default_font).pack(pady=(20, 10))
        entry = ttk.Entry(dialog, width=45, font=self.default_font)
        entry.pack(pady=10)
        entry.focus()

        def on_add():
            name = entry.get().strip()
            if name:
                if self.add_employee(name):
                    self.update_employee_comboboxes()
                    dialog.destroy()
            else:
                messagebox.showwarning("Предупреждение", "Имя сотрудника не может быть пустым.", parent=dialog)

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="Добавить", command=on_add, style='Small.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Отмена", command=dialog.destroy, style='Small.TButton').pack(side='left', padx=5)
        dialog.wait_window()  # Ждем закрытия диалога

    # =============== ВКЛАДКА: ПОИСК ОБОРУДОВАНИЯ ===============
    def create_search_tab(self):
        ttk.Label(self.search_frame, text="Поиск:", font=self.default_font).grid(row=0, column=0, sticky='w', padx=10,
                                                                                 pady=5)
        self.search_entry = ttk.Entry(self.search_frame, width=40, font=self.default_font)
        self.search_entry.grid(row=0, column=1, padx=10, pady=5, sticky='we')
        self.bind_clipboard_events(self.search_entry)
        self.search_entry.bind('<KeyRelease>', self.perform_search)
        ttk.Label(self.search_frame, text="Сотрудник:", font=self.default_font).grid(row=0, column=2, sticky='w',
                                                                                    padx=(20, 10), pady=5)
        self.search_employee_var = tk.StringVar()
        self.search_employee_combo = ttk.Combobox(self.search_frame, textvariable=self.search_employee_var,
                                                  values=[""] + sorted(self.employees_list), width=25,
                                                  font=self.default_font)
        self.search_employee_combo.grid(row=0, column=3, padx=10, pady=5, sticky='we')
        self.search_employee_combo.bind('<<ComboboxSelected>>', self.perform_search)
        clear_button = ttk.Button(self.search_frame, text="Очистить", command=self.clear_search, style='Small.TButton')
        clear_button.grid(row=0, column=4, padx=10, pady=5)
        columns = ("Тип", "Модель", "Серийный номер", "Закрепление", "Дата", "Комментарии")
        self.search_tree = ttk.Treeview(self.search_frame, columns=columns, show='headings', height=15)
        for col in columns:
            self.search_tree.heading(col, text=col,
                                     command=lambda _col=col: self.treeview_sort_column(self.search_tree, _col, False))
            self.search_tree.column(col, width=150, anchor='center')
        self.search_tree.bind('<Double-1>', self.on_tree_double_click)
        self.search_tree.bind('<Button-3>', self.show_context_menu)
        scrollbar = ttk.Scrollbar(self.search_frame, orient="vertical", command=self.search_tree.yview)
        self.search_tree.configure(yscrollcommand=scrollbar.set)
        self.search_tree.grid(row=1, column=0, columnspan=5, padx=10, pady=5, sticky='nsew')
        scrollbar.grid(row=1, column=5, sticky='ns', pady=5)
        export_pdf_btn = ttk.Button(self.search_frame, text="📄 Экспорт результатов поиска в PDF",
                                    command=self.export_search_results_to_pdf, style='Small.TButton')
        export_pdf_btn.grid(row=2, column=0, columnspan=5, pady=10, sticky='we')
        self.search_context_menu = tk.Menu(self.search_tree, tearoff=0)
        self.search_context_menu.add_command(label="Удалить запись", command=self.delete_selected_item)
        self.search_frame.columnconfigure(1, weight=1)
        self.search_frame.columnconfigure(3, weight=1)
        self.search_frame.rowconfigure(1, weight=1)

    # =============== ВКЛАДКА: СОТРУДНИКИ ===============
    def create_employee_tab(self):
        ttk.Label(self.employee_frame, text="Выберите сотрудника:", font=self.default_font).grid(row=0, column=0,
                                                                                                 sticky='w', padx=10,
                                                                                                 pady=5)
        # Расширяем комбобокс на всю ширину
        self.employee_var = tk.StringVar()
        self.employee_combo = ttk.Combobox(self.employee_frame, textvariable=self.employee_var,
                                           values=[""] + sorted(self.employees_list), width=50, font=self.default_font)
        self.employee_combo.grid(row=0, column=1, columnspan=3, padx=10, pady=5, sticky='we')
        self.bind_clipboard_events(self.employee_combo)
        self.employee_combo.bind('<<ComboboxSelected>>', self.show_employee_equipment)
        # --- НОВОЕ ПОЛЕ ПОИСКА СОТРУДНИКОВ ---
        ttk.Label(self.employee_frame, text="Поиск сотрудника:", font=self.default_font).grid(row=1, column=0, sticky='w', padx=10, pady=5)
        self.employee_search_var = tk.StringVar()
        self.employee_search_entry = ttk.Entry(self.employee_frame, textvariable=self.employee_search_var, width=50, font=self.default_font)
        self.employee_search_entry.grid(row=1, column=1, columnspan=3, padx=10, pady=5, sticky='we')
        self.employee_search_entry.bind('<KeyRelease>', self.filter_employees_by_search)
        # Кнопки под комбобоксом — в горизонтальном фрейме
        btn_frame_employees = ttk.Frame(self.employee_frame)
        btn_frame_employees.grid(row=2, column=0, columnspan=4, padx=10, pady=10, sticky='w')
        add_employee_btn = ttk.Button(btn_frame_employees, text="➕ Добавить сотрудника",
                                      command=self.open_add_employee_dialog, style='Small.TButton')
        add_employee_btn.pack(side='left', padx=(0, 10))
        load_from_base_btn = ttk.Button(btn_frame_employees, text="📥 Из базы",
                                        command=self.load_employees_from_inventory, style='Small.TButton')
        load_from_base_btn.pack(side='left', padx=10)
        refresh_button = ttk.Button(btn_frame_employees, text="🔄 Обновить",
                                    command=self.refresh_employee_list, style='Small.TButton')
        refresh_button.pack(side='left', padx=10)
        columns = ("Тип", "Модель", "Серийный номер", "Дата", "Комментарии")
        self.employee_tree = ttk.Treeview(self.employee_frame, columns=columns, show='headings', height=15)
        for col in columns:
            self.employee_tree.heading(col, text=col,
                                       command=lambda _col=col: self.treeview_sort_column(self.employee_tree, _col,
                                                                                          False))
            self.employee_tree.column(col, width=150, anchor='center')
        self.employee_tree.bind('<Double-1>', self.on_tree_double_click)
        self.employee_tree.bind('<Button-3>', self.show_context_menu)
        scrollbar = ttk.Scrollbar(self.employee_frame, orient="vertical", command=self.employee_tree.yview)
        self.employee_tree.configure(yscrollcommand=scrollbar.set)
        self.employee_tree.grid(row=3, column=0, columnspan=4, padx=10, pady=5, sticky='nsew')
        scrollbar.grid(row=3, column=4, sticky='ns', pady=5)
        export_pdf_btn = ttk.Button(self.employee_frame, text="📄 Экспорт в PDF",
                                    command=self.export_employee_results_to_pdf, style='Small.TButton')
        export_pdf_btn.grid(row=4, column=0, columnspan=4, pady=10, sticky='we')
        self.employee_context_menu = tk.Menu(self.employee_tree, tearoff=0)
        self.employee_context_menu.add_command(label="Удалить запись", command=self.delete_selected_item)
        self.employee_frame.columnconfigure(1, weight=1)
        self.employee_frame.rowconfigure(3, weight=1)
        # Инициализация
        self.refresh_employee_list()

    def filter_employees_by_search(self, event=None):
        """Фильтрует список сотрудников в Combobox по введённому тексту"""
        search_term = self.employee_search_var.get().lower().strip()
        filtered_employees = [emp for emp in [""] + sorted(self.employees_list) if search_term in emp.lower()]
        self.employee_combo['values'] = filtered_employees
        if filtered_employees:
            self.employee_combo.current(0)  # Выбираем первый элемент
        else:
            self.employee_combo.set('')
        # Принудительно обновляем отображение оборудования
        self.show_employee_equipment()

    def load_employees_from_inventory(self):
        """Загружает уникальные имена сотрудников из inventory.json в список сотрудников"""
        if not self.inventory_data:
            messagebox.showwarning("Предупреждение", "Нет данных в базе инвентаризации.")
            return
        new_employees = set()
        for item in self.inventory_data:
            assignment = item.get('assignment', '').strip()
            if assignment:
                new_employees.add(assignment)
        if not new_employees:
            messagebox.showinfo("Информация", "В базе нет записей с закреплёнными сотрудниками.")
            return
        added_count = 0
        for emp in new_employees:
            if emp not in self.employees_list:
                self.employees_list.append(emp)
                added_count += 1
        if self.save_employees(self.employees_list):
            self.update_employee_comboboxes()
            messagebox.showinfo("Успех", f"Из базы загружено {added_count} новых сотрудников.\nВсего сотрудников: {len(self.employees_list)}")
        else:
            messagebox.showerror("Ошибка", "Не удалось сохранить список сотрудников.")

    def refresh_employee_list(self):
        """Обновляет список сотрудников в Combobox на основе sotrudniki.json"""
        self.update_employee_comboboxes()

    # =============== ВКЛАДКА: ОБОРУДОВАНИЕ ===============
    def create_equipment_tab(self):
        frame = self.equipment_frame
        ttk.Label(frame, text="Тип оборудования:", font=self.default_font).grid(row=0, column=0, sticky='w', padx=10,
                                                                                pady=5)
        self.equipment_type_entry = ttk.Entry(frame, width=40, font=self.default_font)
        self.equipment_type_entry.grid(row=0, column=1, padx=10, pady=5, sticky='we')
        self.bind_clipboard_events(self.equipment_type_entry)
        add_btn = ttk.Button(frame, text="➕ Добавить тип", command=self.add_equipment_type, style='Small.TButton')
        add_btn.grid(row=0, column=2, padx=10, pady=5)
        columns = ("Тип оборудования",)
        self.equipment_tree = ttk.Treeview(frame, columns=columns, show='headings', height=15)
        self.equipment_tree.heading("Тип оборудования", text="Тип оборудования")
        self.equipment_tree.column("Тип оборудования", width=300, anchor='center')
        self.equipment_tree.bind('<Button-3>', self.show_equipment_context_menu)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.equipment_tree.yview)
        self.equipment_tree.configure(yscrollcommand=scrollbar.set)
        self.equipment_tree.grid(row=1, column=0, columnspan=3, padx=10, pady=5, sticky='nsew')
        scrollbar.grid(row=1, column=3, sticky='ns', pady=5)
        self.equipment_context_menu = tk.Menu(self.equipment_tree, tearoff=0)
        self.equipment_context_menu.add_command(label="🗑️ Удалить тип", command=self.delete_equipment_type)
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(1, weight=1)
        self.refresh_equipment_list()

    def refresh_equipment_list(self):
        for item in self.equipment_tree.get_children():
            self.equipment_tree.delete(item)
        for eq_type in sorted(self.equipment_types):
            self.equipment_tree.insert("", "end", values=(eq_type,))

    def add_equipment_type(self):
        new_type = self.equipment_type_entry.get().strip()
        if not new_type:
            messagebox.showwarning("Предупреждение", "Введите название типа оборудования")
            return
        if new_type in self.equipment_types:
            messagebox.showwarning("Предупреждение", "Такой тип уже существует")
            return
        self.equipment_types.append(new_type)
        if self.save_equipment_types(self.equipment_types):
            messagebox.showinfo("Успех", "Тип оборудования добавлен")
            self.equipment_type_entry.delete(0, tk.END)
            self.refresh_equipment_list()
            self.update_combo_from_data(self.equipment_type_var, 'equipment_type', self.equipment_type_combo)

    def show_equipment_context_menu(self, event):
        item = self.equipment_tree.identify_row(event.y)
        if item:
            self.equipment_tree.selection_set(item)
            self.equipment_context_menu.post(event.x_root, event.y_root)

    def delete_equipment_type(self):
        selected_items = self.equipment_tree.selection()
        if not selected_items:
            messagebox.showwarning("Предупреждение", "Выберите тип для удаления")
            return
        selected_type = self.equipment_tree.item(selected_items[0], 'values')[0]
        in_use = any(item.get('equipment_type') == selected_type for item in self.inventory_data)
        if in_use:
            messagebox.showerror("Ошибка",
                                 f"Тип '{selected_type}' используется в {in_use} записях. Удаление запрещено.")
            return
        if messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить тип '{selected_type}'?"):
            self.equipment_types.remove(selected_type)
            if self.save_equipment_types(self.equipment_types):
                messagebox.showinfo("Успех", "Тип оборудования удален")
                self.refresh_equipment_list()
                self.update_combo_from_data(self.equipment_type_var, 'equipment_type', self.equipment_type_combo)

    # =============== ВКЛАДКА: НАСТРОЙКИ ===============
    def create_settings_tab(self):
        frame = self.settings_frame
        ttk.Label(frame, text="Текущий файл базы:", font=self.default_font).pack(pady=10)
        self.current_path_label = ttk.Label(frame, text=self.inventory_file, font=self.default_font, wraplength=800)
        self.current_path_label.pack(pady=5)
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=10)
        change_file_btn = ttk.Button(btn_frame, text="📂 Выбрать файл inventory.json",
                                     command=self.change_inventory_file, style='Small.TButton')
        change_file_btn.pack(side='left', padx=5)
        change_dir_btn = ttk.Button(btn_frame, text="📁 Выбрать каталог", command=self.choose_inventory_directory,
                                    style='Small.TButton')
        change_dir_btn.pack(side='left', padx=5)
        ttk.Label(frame, text="⚠️ После смены каталога файл будет создан/перезаписан в нём", font=self.default_font,
                  foreground="red").pack(pady=10)

    def choose_inventory_directory(self):
        """Позволяет пользователю выбрать каталог, где будет лежать inventory.json"""
        directory = filedialog.askdirectory(title="Выберите каталог для хранения inventory.json")
        if not directory:
            return
        # Формируем путь к файлу в выбранном каталоге
        new_file = os.path.join(directory, "inventory.json")
        self.inventory_file = new_file
        self.current_path_label.config(text=self.inventory_file)
        # Загружаем или создаём файл
        self.inventory_data = self.load_data()
        self.show_all_data()
        self.refresh_employee_list()
        self.update_history_combobox()
        self.update_serial_combobox()
        messagebox.showinfo("Успех", f"Каталог изменён:\n{directory}\nФайл: {self.inventory_file}")

    def change_inventory_file(self):
        """Позволяет выбрать конкретный файл inventory.json"""
        new_file = filedialog.askopenfilename(
            title="Выберите файл инвентаризации (inventory.json)",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if not new_file:
            return
        try:
            with open(new_file, 'r', encoding='utf-8') as f:
                test_data = json.load(f)
                if not isinstance(test_data, list):
                    raise ValueError("Файл должен содержать массив объектов")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Невозможно загрузить файл:\n{e}")
            return
        self.inventory_file = new_file
        self.inventory_data = self.load_data()
        self.current_path_label.config(text=self.inventory_file)
        self.show_all_data()
        self.refresh_employee_list()
        self.update_history_combobox()
        self.update_serial_combobox()
        messagebox.showinfo("Успех", "Файл базы успешно изменен!")

    # =============== ВКЛАДКА: ИСТОРИЯ ===============
    def create_history_tab(self):
        frame = self.history_frame
        search_frame = ttk.Frame(frame)
        search_frame.pack(pady=10, fill='x', padx=10)
        ttk.Label(search_frame, text="Тип оборудования:", font=self.default_font).grid(row=0, column=0, sticky='w',
                                                                                       padx=(0, 10))
        self.history_type_var = tk.StringVar()
        self.history_type_combo = ttk.Combobox(search_frame, textvariable=self.history_type_var,
                                               values=[""] + sorted(set(
                                                   item.get('equipment_type', '') for item in self.inventory_data if
                                                   item.get('equipment_type'))),
                                               width=25, font=self.default_font)
        self.history_type_combo.grid(row=0, column=1, padx=(0, 10), sticky='we')
        self.history_type_combo.bind('<<ComboboxSelected>>', self.update_serial_combobox)

        ttk.Label(search_frame, text="Серийный номер:", font=self.default_font).grid(row=0, column=2, sticky='w',
                                                                                     padx=(10, 10))
        # УДАЛЕННЫЙ КОД: поле ввода и связанные переменные
        self.history_serial_combo_var = tk.StringVar()
        self.history_serial_combo = ttk.Combobox(search_frame, textvariable=self.history_serial_combo_var,
                                                 values=[], width=25, font=self.default_font)
        self.history_serial_combo.grid(row=0, column=3, padx=(0, 10), sticky='we')
        self.history_serial_combo.bind('<<ComboboxSelected>>', self.show_history_for_equipment)

        refresh_btn = ttk.Button(search_frame, text="🔄 Обновить", command=self.update_serial_combobox,
                                 style='Small.TButton')
        refresh_btn.grid(row=0, column=4, padx=(10, 0))

        search_frame.columnconfigure(1, weight=1)
        search_frame.columnconfigure(3, weight=1)

        # --- ОБНОВЛЁННЫЙ СПИСОК СТОЛБЦОВ: ДОБАВЛЕН "Тип оборудования" ---
        columns = ("Тип оборудования", "Серийный номер", "Сотрудник", "Дата закрепления")
        self.history_tree = ttk.Treeview(frame, columns=columns, show='headings', height=15)
        self.history_tree.heading("Тип оборудования", text="Тип оборудования")
        self.history_tree.heading("Серийный номер", text="Серийный номер")
        self.history_tree.heading("Сотрудник", text="Сотрудник")
        self.history_tree.heading("Дата закрепления", text="Дата закрепления")
        self.history_tree.column("Тип оборудования", width=180, anchor='center')
        self.history_tree.column("Серийный номер", width=150, anchor='center')
        self.history_tree.column("Сотрудник", width=200, anchor='center')
        self.history_tree.column("Дата закрепления", width=150, anchor='center')
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)
        self.history_tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        scrollbar.pack(side='right', fill='y')
        export_hist_btn = ttk.Button(frame, text="📥 Экспорт истории в Excel",
                                     command=self.export_history_to_excel, style='Small.TButton')
        export_hist_btn.pack(pady=10)
        self.update_history_combobox()

    # УДАЛЕН метод filter_serial_numbers_by_input

    def get_filtered_serial_numbers(self):
        """Возвращает отфильтрованный список серийных номеров в зависимости от выбранного типа оборудования"""
        selected_type = self.history_type_var.get()
        if not selected_type:
            serials = sorted(
                set(item.get('serial_number', '') for item in self.inventory_data if item.get('serial_number')))
        else:
            serials = sorted(set(item.get('serial_number', '') for item in self.inventory_data
                                 if item.get('equipment_type') == selected_type and item.get('serial_number')))
        return serials

    def update_serial_combobox(self, event=None):
        """Обновляет список серийных номеров в зависимости от выбранного типа оборудования"""
        serials = self.get_filtered_serial_numbers()
        self.history_serial_combo['values'] = serials
        if serials:
            self.history_serial_combo_var.set(serials[0])
        else:
            self.history_serial_combo_var.set('')
        self.show_history_for_equipment()

    def show_history_for_equipment(self, event=None):
        # Теперь используем только значение из выпадающего списка
        serial = self.history_serial_combo_var.get().strip()
        if not serial:
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)
            return
        # Находим тип оборудования по серийному номеру
        eq_type = "-"
        for item in self.inventory_data:
            if item.get('serial_number') == serial:
                eq_type = item.get('equipment_type', '-')
                break
        history_list = self.get_history_for_equipment(serial)
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        for record in history_list:
            self.history_tree.insert("", "end", values=(
                eq_type,
                serial,
                record["assignment"],
                record["date"]
            ))

    def export_history_to_excel(self):
        """Экспортирует всю историю закреплений в Excel с типом оборудования и автоподбором ширины столбцов"""
        if not self.history_data:
            messagebox.showwarning("Предупреждение", "История пуста")
            return
        try:
            import pandas as pd
            rows = []
            for serial, records in self.history_data.items():
                # Находим тип оборудования для этого серийного номера
                eq_type = "-"
                for item in self.inventory_data:
                    if item.get('serial_number') == serial:
                        eq_type = item.get('equipment_type', '-')
                        break
                for record in records:
                    rows.append({
                        "Тип оборудования": eq_type,
                        "Серийный номер": serial,
                        "Сотрудник": record["assignment"],
                        "Дата закрепления": record["date"]
                    })
            df = pd.DataFrame(rows)
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Сохранить историю в Excel",
                initialdir=os.path.dirname(self.inventory_file)
            )
            if not filename:
                return
            df.to_excel(filename, index=False, engine='openpyxl')
            # --- АВТОПОДБОР ШИРИНЫ СТОЛБЦОВ ---
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            ws = wb.active
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            wb.save(filename)
            webbrowser.open(filename)
            messagebox.showinfo("Успех", f"История экспортирована:\n{filename}")
        except ImportError:
            messagebox.showerror("Ошибка",
                                 "Отсутствует модуль openpyxl. Установите его через pip:\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать историю в Excel:\n{e}")

    # =============== ВКЛАДКА: ИНФО ===============
    def create_about_tab(self):
        center_frame = ttk.Frame(self.about_frame)
        center_frame.pack(expand=True, fill='both')
        info_text = """
        Система инвентаризации оборудования
        Версия: 0.9 (обновлённая)
        Разработано: Разин Григорий
        Контактная информация:
        Email: lantester35@gmail.com
        Функционал:
        - Ведение учета оборудования
        - Поиск и фильтрация данных
        - Экспорт отчетов в PDF и Excel
        - Управление закреплением за сотрудниками
        - Управление типами оборудования
        - Настройка пути к файлу базы
        - Удаление записей (правый клик или кнопка удаления)
        - График распределения типов оборудования
        - История закреплений оборудования
        - Автосохранение каждые 5 минут
        """
        about_text = scrolledtext.ScrolledText(center_frame, width=60, height=15,
                                               font=self.default_font, wrap=tk.WORD)
        about_text.insert('1.0', info_text)
        about_text.config(state='disabled')
        about_text.pack(pady=20, padx=20, expand=True)

    # =============== ОБЩИЕ МЕТОДЫ ===============
    def treeview_sort_column(self, tree, col, reverse):
        """Сортирует Treeview по колонке (только отображение!)"""
        date_columns = ['Дата']
        int_columns = []
        def sort_key(val):
            if col in date_columns:
                try:
                    return datetime.strptime(val, "%d.%m.%Y")
                except:
                    return datetime.min
            elif col in int_columns:
                try:
                    return int(val)
                except:
                    return -1
            else:
                return val.lower() if isinstance(val, str) else val
        data = [(tree.set(k, col), k) for k in tree.get_children('')]
        data.sort(key=lambda t: sort_key(t[0]), reverse=reverse)
        for index, (_, k) in enumerate(data):
            tree.move(k, '', index)
        tree.heading(col, command=lambda: self.treeview_sort_column(tree, col, not reverse))

    def show_context_menu(self, event):
        tree = event.widget
        item = tree.identify_row(event.y)
        if item:
            tree.selection_set(item)
            if tree == self.search_tree:
                self.search_context_menu.post(event.x_root, event.y_root)
            elif tree == self.employee_tree:
                self.employee_context_menu.post(event.x_root, event.y_root)
            elif tree == self.all_tree:
                self.all_context_menu.post(event.x_root, event.y_root)

    def delete_selected_item(self):
        current_tab = self.notebook.index(self.notebook.select())
        if current_tab == 1:  # Поиск
            tree = self.search_tree
        elif current_tab == 2:  # Сотрудники
            tree = self.employee_tree
        elif current_tab == 0:  # Все
            tree = self.all_tree
        else:
            return
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("Предупреждение", "Выберите запись для удаления")
            return
        if not messagebox.askyesno("Подтверждение", "Вы уверены, что хотите удалить выбранную запись?"):
            return
        for item in selected_items:
            values = tree.item(item, 'values')
            serial_number = values[2] if len(values) > 2 else None
            if not serial_number:
                continue
            # Удаляем из self.inventory_data
            for i, inv_item in enumerate(self.inventory_data):
                if inv_item.get('serial_number') == serial_number:
                    del self.inventory_data[i]
                    tree.delete(item)
                    break
            else:
                continue  # Не нашли — пропускаем
        if self.save_data():
            messagebox.showinfo("Успех", "Запись успешно удалена")
            self.refresh_employee_list()
            self.show_all_data()
            self.update_history_combobox()
            self.update_serial_combobox()  # Обновляем комбобоксы в истории

    def add_equipment(self):
        equipment_data = {}
        for field_name, entry in self.entries.items():
            if field_name == "comments":
                equipment_data[field_name] = entry.get("1.0", tk.END).strip()
            else:
                equipment_data[field_name] = entry.get().strip()
        if not equipment_data.get('equipment_type') or not equipment_data.get('serial_number'):
            messagebox.showwarning("Предупреждение", "Заполните обязательные поля: Тип оборудования и Серийный номер")
            return
        if equipment_data['equipment_type'] not in self.equipment_types:
            messagebox.showwarning("Предупреждение", "Выбранный тип оборудования не существует в списке.")
            return
        equipment_data['created_datetime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.inventory_data.append(equipment_data)
        self.add_to_history(equipment_data['serial_number'], equipment_data['assignment'], equipment_data['date'])
        if self.save_data():
            messagebox.showinfo("Успех", "Оборудование успешно добавлено!")
            self.clear_entries()
            self.refresh_employee_list()
            self.show_all_data()
            self.update_history_combobox()
            self.update_serial_combobox()  # Обновляем комбобоксы в истории

    def clear_entries(self):
        for field_name, entry in self.entries.items():
            if field_name == "comments":
                entry.delete("1.0", tk.END)
            elif field_name == "date":
                entry.delete(0, tk.END)
                entry.insert(0, datetime.now().strftime("%d.%m.%Y"))
            else:
                entry.delete(0, tk.END)
        if 'equipment_type' in self.entries:
            self.entries['equipment_type'].set('')

    def perform_search(self, event=None):
        search_text = self.search_entry.get().lower().strip()
        # Получаем выбранный сотрудник (если есть)
        selected_employee = self.search_employee_var.get().strip()
        for item in self.search_tree.get_children():
            self.search_tree.delete(item)
        if not search_text and not selected_employee:
            return
        for item in self.inventory_data:
            # Проверяем соответствие поисковому запросу
            matches_search = any(search_text in str(value).lower() for value in item.values() if value)
            # Проверяем соответствие фильтру по сотруднику
            matches_employee = (not selected_employee) or (item.get('assignment', '') == selected_employee)
            if matches_search and matches_employee:
                # Найдём индекс элемента в inventory_data
                idx = None
                for i, inv_item in enumerate(self.inventory_data):
                    if inv_item.get('serial_number') == item.get('serial_number'):
                        idx = i
                        break
                self.search_tree.insert("", "end", values=(
                    item.get('equipment_type', ''),
                    item.get('model', ''),
                    item.get('serial_number', ''),
                    item.get('assignment', ''),
                    item.get('date', ''),
                    (item.get('comments', '')[:50] + '...') if item.get('comments') and len(
                        item.get('comments')) > 50 else item.get('comments', '')
                ), tags=(str(idx),))  # 👈 ДОБАВЛЕНО! ТЕГИ ДЛЯ РЕДАКТИРОВАНИЯ

    def clear_search(self):
        self.search_entry.delete(0, tk.END)
        self.search_employee_var.set('')  # Очищаем фильтр по сотруднику
        for item in self.search_tree.get_children():
            self.search_tree.delete(item)

    def refresh_employee_list(self):
        # Теперь мы просто обновляем Combobox, а не формируем список из inventory_data
        self.update_employee_comboboxes()

    def update_history_combobox(self):
        """Обновляет список серийных номеров в комбобоксе истории"""
        serials = sorted(
            set(item.get('serial_number', '') for item in self.inventory_data if item.get('serial_number')))
        self.update_combo_from_data(self.history_serial_combo_var, 'serial_number', self.history_serial_combo)

    def update_combo_from_data(self, var, key_name, combo_widget):
        """Универсальный метод для обновления Combobox на основе списка self.inventory_data"""
        values = sorted(set(item.get(key_name, '') for item in self.inventory_data if item.get(key_name)))
        combo_widget['values'] = values
        if values:
            var.set(values[0])  # Установить первый как значение по умолчанию
        else:
            var.set('')

    def show_employee_equipment(self, event=None):
        employee = self.employee_var.get()
        for item in self.employee_tree.get_children():
            self.employee_tree.delete(item)
        if not employee:
            return
        for item in self.inventory_data:
            if item.get('assignment') == employee:
                self.employee_tree.insert("", "end", values=(
                    item.get('equipment_type', ''),
                    item.get('model', ''),
                    item.get('serial_number', ''),
                    item.get('date', ''),
                    (item.get('comments', '')[:50] + '...') if item.get('comments') and len(
                        item.get('comments')) > 50 else item.get('comments', '')
                ))

    def show_all_data(self):
        for item in self.all_tree.get_children():
            self.all_tree.delete(item)
        self.inventory_data = self.load_data()  # Перегружаем из файла на всякий случай
        for item in self.inventory_data:
            # Сохраняем индекс как tag для точного редактирования
            self.all_tree.insert("", "end", values=(
                item.get('equipment_type', ''),
                item.get('model', ''),
                item.get('serial_number', ''),
                item.get('assignment', ''),
                item.get('date', ''),
                (item.get('comments', '')[:50] + '...') if item.get('comments') and len(
                    item.get('comments')) > 50 else item.get('comments', '')
            ), tags=(str(self.inventory_data.index(item)),))
        self.refresh_employee_list()
        self.update_history_combobox()
        self.update_serial_combobox()  # Обновляем комбобокс типа в истории

    def on_tree_double_click(self, event):
        tree = event.widget
        item = tree.identify('item', event.x, event.y)
        column = tree.identify_column(event.x)
        if not item or item not in tree.get_children():
            return
        col_index = int(column.replace('#', '')) - 1
        current_values = tree.item(item, 'values')
        current_value = current_values[col_index]
        # Определяем поля в зависимости от таблицы
        if tree == self.employee_tree:
            field_names = {
                0: 'equipment_type',
                1: 'model',
                2: 'serial_number',
                3: 'date',
                4: 'comments'
            }
        elif tree == self.all_tree:
            field_names = {
                0: 'equipment_type',
                1: 'model',
                2: 'serial_number',
                3: 'assignment',
                4: 'date',
                5: 'comments'
            }
        elif tree == self.search_tree:  # 👈 ДОБАВЛЕНО!
            field_names = {
                0: 'equipment_type',
                1: 'model',
                2: 'serial_number',
                3: 'assignment',
                4: 'date',
                5: 'comments'
            }
        else:
            return
        field_name = field_names.get(col_index)
        if not field_name:
            return
        # Получаем индекс записи в self.inventory_data
        tags = tree.item(item, 'tags')
        if not tags:
            return
        try:
            idx = int(tags[0])
            if idx < 0 or idx >= len(self.inventory_data):
                return
        except:
            return
        self.edit_cell(tree, item, col_index, field_name, current_value, idx)

    def edit_cell(self, tree, item, col_index, field_name, current_value, data_index):
        bbox = tree.bbox(item, column=f'#{col_index + 1}')
        if not bbox:  # Элемент может быть не виден
            return
        if field_name == 'comments':
            text_edit = scrolledtext.ScrolledText(tree, width=40, height=4, font=self.default_font)
            text_edit.insert('1.0', current_value)
            text_edit.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3] * 3)
            text_edit.focus()
            self.bind_clipboard_events(text_edit)

            def save_edit(event=None):
                new_value = text_edit.get('1.0', tk.END).strip()
                text_edit.destroy()
                current_values = list(tree.item(item, 'values'))
                current_values[col_index] = new_value
                tree.item(item, values=current_values)
                self.inventory_data[data_index][field_name] = new_value
                self.save_data()

            def cancel_edit(event=None):
                text_edit.destroy()

            text_edit.bind('<Return>', save_edit)
            text_edit.bind('<Escape>', cancel_edit)
            text_edit.bind('<FocusOut>', lambda e: save_edit())
        elif field_name == 'assignment':  # 👈 ОБРАБОТКА ЧЕРЕЗ COMBOBOX ВО ВСЕХ ТАБЛИЦАХ
            # Создаём Combobox с доступными сотрудниками
            combo_edit = ttk.Combobox(
                tree,
                values=[""] + sorted(self.employees_list),
                width=bbox[2] // 8,
                font=self.default_font,
                state='readonly'  # Только выбор, без ввода
            )
            combo_edit.set(current_value)  # Устанавливаем текущее значение
            combo_edit.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            combo_edit.focus()

            def save_edit(event=None):
                new_value = combo_edit.get().strip()
                combo_edit.destroy()
                current_values = list(tree.item(item, 'values'))
                current_values[col_index] = new_value
                tree.item(item, values=current_values)
                old_value = current_value  # Сохраняем старое значение до изменения
                self.inventory_data[data_index][field_name] = new_value
                # Если закрепление изменилось — записываем в историю
                serial_number = self.inventory_data[data_index].get('serial_number', '')
                if serial_number and old_value != new_value:
                    current_date = datetime.now().strftime("%d.%m.%Y")
                    self.add_to_history(serial_number, new_value, current_date)
                    print(f"[HISTORY CHANGE] Serial: {serial_number}, Changed from '{old_value}' to '{new_value}' on {current_date}")
                self.save_data()

            def cancel_edit(event=None):
                combo_edit.destroy()

            combo_edit.bind('<Return>', save_edit)
            combo_edit.bind('<Escape>', cancel_edit)
            combo_edit.bind('<FocusOut>', lambda e: save_edit())
            combo_edit.bind('<<ComboboxSelected>>', lambda e: save_edit())  # Сохраняем при выборе
        else:  # Все остальные поля — обычный Entry
            entry_edit = ttk.Entry(tree, width=bbox[2] // 8, font=self.default_font)
            entry_edit.insert(0, current_value)
            entry_edit.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            entry_edit.focus()
            self.bind_clipboard_events(entry_edit)

            def save_edit(event=None):
                new_value = entry_edit.get().strip()
                entry_edit.destroy()
                current_values = list(tree.item(item, 'values'))
                current_values[col_index] = new_value
                tree.item(item, values=current_values)
                self.inventory_data[data_index][field_name] = new_value
                self.save_data()

            def cancel_edit(event=None):
                entry_edit.destroy()

            entry_edit.bind('<Return>', save_edit)
            entry_edit.bind('<Escape>', cancel_edit)
            entry_edit.bind('<FocusOut>', lambda e: save_edit())

    # =============== ЭКСПОРТ В EXCEL ===============
    def export_to_excel(self):
        """Экспорт полного списка оборудования в Excel"""
        if not self.inventory_data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        try:
            # Импорт здесь, чтобы не грузить при старте
            import pandas as pd
            # Получаем текущие отсортированные данные из Treeview
            current_data = []
            # Сначала получаем данные из Treeview, если она активна, иначе из self.inventory_data
            active_tab = self.notebook.index(self.notebook.select())
            if active_tab == 0:  # Показать всё
                # Получаем все строки из Treeview
                for item in self.all_tree.get_children():
                    values = self.all_tree.item(item, 'values')
                    # Преобразуем обратно в словарь
                    item_dict = {
                        'equipment_type': values[0],
                        'model': values[1],
                        'serial_number': values[2],
                        'assignment': values[3],
                        'date': values[4],
                        'comments': values[5]
                    }
                    current_data.append(item_dict)
                # Используем отсортированные данные
                data_to_export = current_data
            else:
                # Используем оригинальные данные
                data_to_export = self.inventory_data
            df = pd.DataFrame(data_to_export)
            df = df.rename(columns={
                'equipment_type': 'Тип',
                'model': 'Модель',
                'serial_number': 'Серийный номер',
                'assignment': 'Закрепление',
                'date': 'Дата',
                'comments': 'Комментарии'
            })
            desired_order = ['Тип', 'Модель', 'Серийный номер', 'Закрепление', 'Дата', 'Комментарии']
            df = df[desired_order]
            # Преобразуем столбец 'Дата' в datetime для корректного отображения в Excel
            df['Дата'] = pd.to_datetime(df['Дата'], format='%d.%m.%Y', errors='coerce')
            # Открываем диалог сохранения файла
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Сохранить отчет в Excel",
                initialdir=os.path.dirname(self.inventory_file)  # Начальный каталог - каталог базы
            )
            if not filename:  # Пользователь нажал "Отмена"
                return
            df.to_excel(filename, index=False, engine='openpyxl')
            # Автоматическая настройка ширины столбцов в Excel
            # Для этого мы используем openpyxl напрямую после создания файла
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            ws = wb.active
            # Автоматическое подгонение ширины столбцов
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            wb.save(filename)
            webbrowser.open(filename)
            messagebox.showinfo("Успех", f"Отчет сохранен:\n{filename}")
        except ImportError:
            messagebox.showerror("Ошибка",
                                 "Отсутствует модуль openpyxl. Установите его через pip:\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать Excel отчет: {e}")

    # =============== ГРАФИК ===============
    def show_equipment_graph(self):
        """Отображает график распределения типов оборудования"""
        if not self.inventory_data:
            messagebox.showwarning("Предупреждение", "Нет данных для построения графика")
            return
        try:
            # Импорт здесь, чтобы не грузить при старте
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            type_counts = {}
            for item in self.inventory_data:
                eq_type = item.get('equipment_type', 'Не указано')
                type_counts[eq_type] = type_counts.get(eq_type, 0) + 1
            sorted_types = sorted(type_counts.items(), key=lambda x: x[1], reverse=True)
            types, counts = zip(*sorted_types) if sorted_types else ([], [])
            graph_window = tk.Toplevel(self.root)
            graph_window.title("График распределения оборудования")
            graph_window.geometry("900x700")
            fig, ax = plt.subplots(figsize=(10, 8))
            bars = ax.bar(types, counts, color='skyblue', edgecolor='black')
            ax.set_title('Количество единиц оборудования по типам', fontsize=16, fontweight='bold')
            ax.set_xlabel('Тип оборудования', fontsize=12)
            ax.set_ylabel('Количество', fontsize=12)
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2., height + 0.1,
                        f'{int(height)}', ha='center', va='bottom', fontsize=10)
            canvas = FigureCanvasTkAgg(fig, master=graph_window)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            close_btn = ttk.Button(graph_window, text="Закрыть", command=graph_window.destroy, style='Big.TButton')
            close_btn.pack(pady=10)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось построить график: {e}")

    # =============== ЭКСПОРТ В PDF ===============
    def export_to_pdf(self):
        if not self.inventory_data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        try:
            # Проверка шрифта
            font_path = self._get_asset_path('ChakraPetch-Regular.ttf')
            if not os.path.exists(font_path):
                font_path = self._get_asset_path('assets/fonts/ChakraPetch-Regular.ttf')
                if not os.path.exists(font_path):
                    messagebox.showerror("Ошибка",
                                         "Отсутствует файл шрифта: ChakraPetch-Regular.ttf\nПроверьте, что файл находится в корне проекта или в папке assets/fonts/")
                    return
            pdf = PDFWithCyrillic(orientation='L')
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            pdf.set_font("ChakraPetch", '', 18)
            pdf.cell(0, 10, "Полный отчет по инвентаризации оборудования", 0, 1, 'C')
            pdf.ln(10)
            pdf.set_font("ChakraPetch", '', 12)
            pdf.cell(0, 10, f"Отчет сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}", 0, 1, 'R')
            pdf.ln(5)
            total_equipment = len(self.inventory_data)
            unique_employees = len(
                set(item.get('assignment', '') for item in self.inventory_data if item.get('assignment')))
            pdf.set_font("ChakraPetch", '', 14)
            pdf.cell(0, 10, "Общая статистика:", 0, 1)
            pdf.cell(0, 10, f"Всего единиц оборудования: {total_equipment}", 0, 1)
            pdf.cell(0, 10, f"Количество сотрудников с оборудованием: {unique_employees}", 0, 1)
            pdf.ln(10)
            columns = ["Тип", "Модель", "Серийный номер", "Закрепление", "Дата", "Комментарии"]
            # Получаем текущие отсортированные данные из Treeview
            data_rows = []
            active_tab = self.notebook.index(self.notebook.select())
            if active_tab == 0:  # Показать всё
                # Получаем все строки из Treeview
                for item in self.all_tree.get_children():
                    values = self.all_tree.item(item, 'values')
                    row = [
                        values[0] or '-',
                        values[1] or '-',
                        values[2] or '-',
                        values[3] or '-',
                        values[4] or '-',
                        (values[5][:50] + '...') if values[5] and len(values[5]) > 50 else (values[5] or '-')
                    ]
                    data_rows.append(row)
            else:
                # Используем оригинальные данные
                for item in self.inventory_data:
                    row = [
                        item.get('equipment_type', '') or '-',
                        item.get('model', '') or '-',
                        item.get('serial_number', '') or '-',
                        item.get('assignment', '') or '-',
                        item.get('date', '') or '-',
                        (item.get('comments', '')[:50] + '...') if item.get('comments') and len(
                            item.get('comments')) > 50 else (item.get('comments', '') or '-')
                    ]
                    data_rows.append(row)
            pdf.set_font("ChakraPetch", '', 12)
            col_widths = []
            for col_index in range(len(columns)):
                max_width = pdf.get_string_width(columns[col_index]) + 6
                for row in data_rows:
                    w = pdf.get_string_width(str(row[col_index])) + 6
                    if w > max_width:
                        max_width = w
                col_widths.append(max_width)
            pdf.set_font("ChakraPetch", '', 14)
            for i, col in enumerate(columns):
                pdf.cell(col_widths[i], 10, col, 1, new_x="RIGHT", new_y="TOP", align='C')
            pdf.ln()
            pdf.set_font("ChakraPetch", '', 12)
            for row in data_rows:
                for i, cell_text in enumerate(row):
                    pdf.cell(col_widths[i], 10, str(cell_text), 1, new_x="RIGHT", new_y="TOP")
                pdf.ln()
            # Открываем диалог сохранения файла
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Сохранить отчет в PDF",
                initialdir=os.path.dirname(self.inventory_file)  # Начальный каталог - каталог базы
            )
            if not filename:  # Пользователь нажал "Отмена"
                return
            pdf.output(filename)
            webbrowser.open(filename)
            messagebox.showinfo("Успех", f"Отчет сохранен:\n{filename}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать PDF отчет: {e}")

    def export_search_results_to_pdf(self):
        items = self.search_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        try:
            # Проверка шрифта
            font_path = self._get_asset_path('ChakraPetch-Regular.ttf')
            if not os.path.exists(font_path):
                font_path = self._get_asset_path('assets/fonts/ChakraPetch-Regular.ttf')
                if not os.path.exists(font_path):
                    messagebox.showerror("Ошибка",
                                         "Отсутствует файл шрифта: ChakraPetch-Regular.ttf\nПроверьте, что файл находится в корне проекта или в папке assets/fonts/")
                    return
            pdf = PDFWithCyrillic(orientation='L')
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            pdf.set_font("ChakraPetch", '', 14)
            pdf.cell(0, 10, "Отчет по результатам поиска оборудования", 0, 1, 'C')
            pdf.ln(10)
            columns = ["Тип", "Модель", "Серийный номер", "Закрепление", "Дата", "Комментарии"]
            data_rows = [self.search_tree.item(item, 'values') for item in items]
            pdf.set_font("ChakraPetch", '', 12)
            col_widths = []
            for col_index in range(len(columns)):
                max_width = pdf.get_string_width(columns[col_index]) + 6
                for row in data_rows:
                    w = pdf.get_string_width(str(row[col_index])) + 6
                    if w > max_width:
                        max_width = w
                col_widths.append(max_width)
            pdf.set_font("ChakraPetch", '', 14)
            for i, col in enumerate(columns):
                pdf.cell(col_widths[i], 10, col, 1, new_x="RIGHT", new_y="TOP", align='C')
            pdf.ln()
            pdf.set_font("ChakraPetch", '', 12)
            for row in data_rows:
                for i, cell_text in enumerate(row):
                    pdf.cell(col_widths[i], 10, str(cell_text), 1, new_x="RIGHT", new_y="TOP")
                pdf.ln()
            # Открываем диалог сохранения файла
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Сохранить отчет в PDF",
                initialdir=os.path.dirname(self.inventory_file)  # Начальный каталог - каталог базы
            )
            if not filename:  # Пользователь нажал "Отмена"
                return
            pdf.output(filename)
            webbrowser.open(filename)
            messagebox.showinfo("Успех", f"Отчет сохранен:\n{filename}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать PDF отчет: {e}")

    def export_employee_results_to_pdf(self):
        items = self.employee_tree.get_children()
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return
        try:
            # Проверка шрифта
            font_path = self._get_asset_path('ChakraPetch-Regular.ttf')
            if not os.path.exists(font_path):
                font_path = self._get_asset_path('assets/fonts/ChakraPetch-Regular.ttf')
                if not os.path.exists(font_path):
                    messagebox.showerror("Ошибка",
                                         "Отсутствует файл шрифта: ChakraPetch-Regular.ttf\nПроверьте, что файл находится в корне проекта или в папке assets/fonts/")
                    return
            pdf = PDFWithCyrillic(orientation='L')
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            employee_name = self.employee_var.get()
            pdf.set_font("ChakraPetch", '', 14)
            pdf.cell(0, 10, f"Отчет по оборудованию сотрудника: {employee_name}", 0, 1, 'C')
            pdf.ln(10)
            columns = ["Тип", "Модель", "Серийный номер", "Дата", "Комментарии"]
            data_rows = [self.employee_tree.item(item, 'values') for item in items]
            pdf.set_font("ChakraPetch", '', 12)
            col_widths = []
            for col_index in range(len(columns)):
                max_width = pdf.get_string_width(columns[col_index]) + 6
                for row in data_rows:
                    w = pdf.get_string_width(str(row[col_index])) + 6
                    if w > max_width:
                        max_width = w
                col_widths.append(max_width)
            pdf.set_font("ChakraPetch", '', 14)
            for i, col in enumerate(columns):
                pdf.cell(col_widths[i], 10, col, 1, new_x="RIGHT", new_y="TOP", align='C')
            pdf.ln()
            pdf.set_font("ChakraPetch", '', 12)
            for row in data_rows:
                for i, cell_text in enumerate(row):
                    pdf.cell(col_widths[i], 10, str(cell_text), 1, new_x="RIGHT", new_y="TOP")
                pdf.ln()
            # Открываем диалог сохранения файла
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Сохранить отчет в PDF",
                initialdir=os.path.dirname(self.inventory_file)  # Начальный каталог - каталог базы
            )
            if not filename:  # Пользователь нажал "Отмена"
                return
            pdf.output(filename)
            webbrowser.open(filename)
            messagebox.showinfo("Успех", f"Отчет сохранен:\n{filename}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать PDF отчет: {e}")

    # =============== АВТОСОХРАНЕНИЕ ===============
    def schedule_auto_save(self):
        """Запускает цикл автосохранения каждые 5 минут"""

        def auto_save():
            if self.save_data():
                print(f"[AUTO-SAVE] Данные сохранены в {datetime.now().strftime('%H:%M:%S')}")
            self.root.after(self.auto_save_interval, auto_save)

        self.root.after(self.auto_save_interval, auto_save)

    def _get_asset_path(self, filename):
        """Получает путь к файлу внутри пакета или в рабочей директории.
           Ищет в порядке: корень проекта → assets/fonts/ → MEIPASS"""
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
            path1 = os.path.join(base_path, 'assets', 'fonts', filename)
            path2 = os.path.join(base_path, filename)
            if os.path.exists(path1):
                return path1
            elif os.path.exists(path2):
                return path2
            else:
                raise FileNotFoundError(f"Шрифт не найден в MEIPASS: {filename}")
        else:
            base_path = os.path.dirname(__file__)
            path1 = os.path.join(base_path, filename)  # Корень проекта
            path2 = os.path.join(base_path, 'assets', 'fonts', filename)  # Подпапка
            if os.path.exists(path1):
                return path1
            elif os.path.exists(path2):
                return path2
            else:
                raise FileNotFoundError(f"Шрифт не найден ни в корне проекта, ни в assets/fonts/: {filename}")


def main():
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()